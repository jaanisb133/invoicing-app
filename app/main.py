"""
V-Rēķini — Multi-tenant SaaS Invoice Manager (FastAPI)
"""

import os
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".env"))

import json
import secrets
import datetime
import asyncio
import logging
import smtplib
import time
from collections import defaultdict
from urllib.parse import quote
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from fastapi import FastAPI, Request, Form, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse, PlainTextResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.base import BaseHTTPMiddleware
from itsdangerous import URLSafeTimedSerializer

logger = logging.getLogger("vrekini")
# Surface our INFO/WARNING/EXCEPTION logs to journalctl. Without this the
# logger has no handler and every logger.info/exception is silently dropped.
if not logger.handlers:
    _h = logging.StreamHandler()
    _h.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(name)s: %(message)s"))
    logger.addHandler(_h)
    logger.setLevel(logging.INFO)
    logger.propagate = False

from app import database as db
from app import registry
from app.pdf_generator import generate_invoice_pdf, TEMPLATES
from app.einvoice import generate_einvoice_xml, generate_einvoice_file

app = FastAPI(title="V-Rēķini")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

# Cache-busting version for style.css. Bump this on every CSS change.
# Exposed as a Jinja global so the standalone templates (landing, login,
# pricing, ...) stay in step with base.html — they used to carry their own
# hardcoded ?v= and silently drifted behind, serving visitors a stale
# stylesheet on exactly the pages new users land on.
CSS_VERSION = "26"
templates.env.globals["css_version"] = CSS_VERSION

OFFLINE_MODE = os.getenv("OFFLINE_MODE", "").lower() in ("1", "true", "yes")

# ---------------------------------------------------------------------------
# Offline license protection (machine-locked)
# ---------------------------------------------------------------------------
import hashlib
import uuid
import platform

# SHA-256 hashes of valid license keys (keys are NOT stored in source code)
_OFFLINE_LICENSE_HASHES = {
    "01e59cba7c8bf7cc942afe8571b458974d92429011fa726e48a59acbdd23e4b2",
    "a55071558393e2523fcc9943ff031d55227905d97cac054be69686f7f7b46bca",
    "0cb81b36d79f902673244912bb1f705967258d1f9b7f85c5c87a4c76bfcd4de6",
}

def _get_machine_id() -> str:
    """Generate a stable fingerprint for this machine."""
    raw = f"{platform.node()}|{uuid.getnode()}"
    return hashlib.sha256(raw.encode()).hexdigest()

def _get_license_file_path():
    data_dir = os.environ.get("VREKINI_DB_PATH", "")
    if data_dir:
        return os.path.join(os.path.dirname(data_dir), "license.key")
    return os.path.join(BASE_DIR, "..", "data", "license.key")

def _is_offline_licensed():
    """Check if a valid license key file exists and is bound to this machine."""
    if not OFFLINE_MODE:
        return True
    path = _get_license_file_path()
    if not os.path.exists(path):
        return False
    try:
        lines = open(path, "r", encoding="utf-8").read().strip().splitlines()
        if len(lines) < 2:
            return False
        key = lines[0].strip().upper()
        stored_machine = lines[1].strip()
        h = hashlib.sha256(key.encode()).hexdigest()
        return h in _OFFLINE_LICENSE_HASHES and stored_machine == _get_machine_id()
    except Exception:
        return False

def _activate_license(key: str) -> bool:
    """Validate and save a license key bound to this machine. Returns True if valid."""
    key = key.strip().upper()
    h = hashlib.sha256(key.encode()).hexdigest()
    if h not in _OFFLINE_LICENSE_HASHES:
        return False
    path = _get_license_file_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"{key}\n{_get_machine_id()}\n")
    return True

UNITS = ["gab", "kg", "kaste", "iepak.", "l", "h", "m", "m²", "m³"]


# --- Simple in-memory rate limiter ---
class RateLimiter:
    """IP-based sliding window rate limiter."""

    def __init__(self):
        self._hits = defaultdict(list)

    def is_allowed(self, key, max_requests, window_seconds):
        """Return True if the request is allowed, False if rate limited."""
        now = time.monotonic()
        hits = self._hits[key]
        # Prune old entries
        cutoff = now - window_seconds
        self._hits[key] = hits = [t for t in hits if t > cutoff]
        if len(hits) >= max_requests:
            return False
        hits.append(now)
        return True


_rate_limiter = RateLimiter()


def _get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"

# --- Centralised email configuration ---
# Emails are sent from a single V-Rēķini address; Reply-To is set to
# the user's own email so clients can reply directly to them.
SMTP_HOST = os.getenv("SMTP_HOST", "server50.areait.lv")
SMTP_PORT = int(os.getenv("SMTP_PORT", "465"))
SMTP_USER = os.getenv("SMTP_USER", "rekini@v-rekini.lv")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_FROM = os.getenv("SMTP_FROM", "V-Rēķini <rekini@v-rekini.lv>")
SMTP_SSL = os.getenv("SMTP_SSL", "true").lower() in ("true", "1", "yes")

# --- Brevo transactional email API ---
BREVO_API_KEY = os.getenv("BREVO_API_KEY", "")
BREVO_SENDER_EMAIL = os.getenv("BREVO_SENDER_EMAIL", "rekini@v-rekini.lv")
BREVO_SENDER_NAME = os.getenv("BREVO_SENDER_NAME", "V-Rēķini")

# --- EveryPay / SEB E-commerce ---
from app import everypay


def _smtp_connect():
    """Return an authenticated SMTP connection (SSL or STARTTLS)."""
    if SMTP_SSL:
        server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=15)
    else:
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15)
        server.starttls()
    server.login(SMTP_USER, SMTP_PASS)
    return server


def _send_email(*, to_email: str, subject: str, body: str,
                reply_to: str = "", attachment_path: str = "", sender_name: str = "",
                attachment_name: str = ""):
    """Send email via SMTP (preferred) or Brevo API fallback.

    attachment_name overrides the name the recipient sees. Needed because the
    e-invoice XML is written to a temp file, so the path's basename is
    something like tmpah03c_ed.xml rather than the invoice number.
    """
    if SMTP_PASS:
        logger.info("send_email: provider=SMTP to=%s subject=%r", to_email, subject[:80])
        return _send_via_smtp(
            to_email=to_email, subject=subject, body=body,
            reply_to=reply_to, attachment_path=attachment_path, sender_name=sender_name,
            attachment_name=attachment_name,
        )
    if BREVO_API_KEY:
        logger.info("send_email: provider=Brevo to=%s subject=%r", to_email, subject[:80])
        return _send_via_brevo(
            to_email=to_email, subject=subject, body=body,
            reply_to=reply_to, attachment_path=attachment_path, sender_name=sender_name,
            attachment_name=attachment_name,
        )
    raise RuntimeError("E-pasta serviss nav konfigurēts (nav ne SMTP_PASS, ne BREVO_API_KEY).")


def _send_via_brevo(*, to_email, subject, body, reply_to="", attachment_path="", sender_name="", attachment_name=""):
    """Send email using Brevo HTTP API."""
    import httpx
    import base64

    payload = {
        "sender": {"name": sender_name or BREVO_SENDER_NAME, "email": BREVO_SENDER_EMAIL},
        "to": [{"email": to_email}],
        "subject": subject,
        "textContent": body,
    }
    if reply_to:
        payload["replyTo"] = {"email": reply_to}

    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            content_b64 = base64.b64encode(f.read()).decode()
        payload["attachment"] = [{
            "content": content_b64,
            "name": attachment_name or os.path.basename(attachment_path),
        }]

    resp = httpx.post(
        "https://api.brevo.com/v3/smtp/email",
        headers={
            "api-key": BREVO_API_KEY,
            "Content-Type": "application/json",
            "Accept": "application/json",
        },
        json=payload,
        timeout=30,
    )
    body_preview = resp.text[:300] if resp.text else ""
    logger.info("brevo_send: status=%s sender=%s to=%s body=%s",
                resp.status_code, BREVO_SENDER_EMAIL, to_email, body_preview)
    if resp.status_code not in (200, 201):
        raise RuntimeError(f"Brevo API kļūda ({resp.status_code}): {resp.text}")


def _send_via_smtp(*, to_email, subject, body, reply_to="", attachment_path="", sender_name="", attachment_name=""):
    """Send email using SMTP."""
    from email.utils import formataddr, parseaddr
    msg = MIMEMultipart()
    # Encode the From display name with RFC 2047 so non-ASCII (Latvian ē, etc.)
    # is handled properly. Gmail's spam filter silently drops messages with raw
    # non-ASCII bytes in headers — formataddr produces "=?utf-8?b?...?=" form.
    if sender_name:
        msg["From"] = formataddr((sender_name, SMTP_USER))
    else:
        # SMTP_FROM may be either "addr@host" or "Display Name <addr@host>".
        parsed_name, parsed_addr = parseaddr(SMTP_FROM)
        if parsed_addr:
            msg["From"] = formataddr((parsed_name, parsed_addr))
        else:
            msg["From"] = SMTP_FROM
    msg["To"] = to_email
    msg["Subject"] = subject
    if reply_to:
        msg["Reply-To"] = reply_to

    msg.attach(MIMEText(body, "plain", "utf-8"))

    if attachment_path and os.path.exists(attachment_path):
        # Derive the type from the file — the attachment is a PDF most of the
        # time but can also be a PEPPOL XML e-invoice, and labelling that as
        # application/pdf makes mail clients refuse to open it.
        import mimetypes
        ctype, _enc = mimetypes.guess_type(attachment_path)
        maintype, _slash, subtype = (ctype or "application/octet-stream").partition("/")
        with open(attachment_path, "rb") as f:
            part = MIMEBase(maintype, subtype or "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment",
                            filename=attachment_name or os.path.basename(attachment_path))
            msg.attach(part)

    with _smtp_connect() as server:
        refused = server.send_message(msg)
        # send_message returns a dict of refused recipients (empty = all accepted).
        # An empty dict here means the SMTP server has accepted the message for delivery
        # but does NOT guarantee inbox delivery — bounces/blocks happen async.
        logger.info("smtp_send: from=%s to=%s refused=%s",
                    msg.get("From"), to_email, refused or "none")


SESSION_COOKIE = "session"
SESSION_MAX_AGE = 60 * 60 * 24 * 30  # 30 days


def _get_serializer():
    secret = db.get_setting("session_secret", "")
    if not secret:
        secret = secrets.token_urlsafe(32)
        db.set_setting("session_secret", secret)
    return URLSafeTimedSerializer(secret)


def _get_current_user(request: Request):
    cookie = request.cookies.get(SESSION_COOKIE)
    if not cookie:
        return None
    try:
        s = _get_serializer()
        user_id = s.loads(cookie, max_age=SESSION_MAX_AGE)
        return db.get_user(user_id)
    except Exception:
        return None


def _set_session_cookie(response, user_id):
    s = _get_serializer()
    token = s.dumps(user_id)
    response.set_cookie(SESSION_COOKIE, token, max_age=SESSION_MAX_AGE,
                        httponly=True, samesite="lax")
    return response


def _stock_enabled(user_id):
    return db.get_user_setting(user_id, "stock_enabled", "0") == "1"


def _user_settings(user_id):
    return db.get_all_user_settings(user_id)


def _base_context(request):
    """Common template context for authenticated pages."""
    user = request.state.user
    tier = user.get("tier", "free")
    return {
        "request": request,
        "current_user": user,
        "stock_enabled": _stock_enabled(user["id"]),
        "tier": tier,
        "tier_label": db.TIER_LIMITS.get(tier, {}).get("label", "Bezmaksas"),
        "tier_limits": db.get_tier_limits(tier),
        "needs_setup": not db.get_user_setting(user["id"], "company_name"),
        "offline_mode": OFFLINE_MODE,
    }


def _with_document_client(clients, uid, doc):
    """Make sure a document's own client is in the picker list.

    get_all_clients() hides deleted and one-time clients, so a document
    attached to one would render with an empty client box even though the
    client is still on the document. The stored client_id is untouched either
    way, but showing a blank field on a document that has a client invites the
    user to pick the wrong one."""
    if not doc or not doc.get("client_id"):
        return clients
    if any(c["id"] == doc["client_id"] for c in clients):
        return clients
    own = db.get_client(doc["client_id"])
    if own and own.get("user_id") == uid:
        return list(clients) + [own]
    return clients


FREE_TEMPLATE = "classic"

# Offer lifecycle. Reuses documents.status, which offers already default to
# 'issued'; the unpaid/overdue queries are all scoped to doc_type='sell', so
# these values never leak into invoice figures.
OFFER_STATUSES = {
    "issued": "Gaida atbildi",
    "accepted": "Pieņemts",
    "rejected": "Noraidīts",
}
OFFER_STATUS_PILL = {
    "issued": "pill-neutral",
    "accepted": "pill-success",
    "rejected": "pill-danger",
}


def _resolve_template(user, template=""):
    """Clamp a requested PDF template to what the user's tier allows.

    The pickers hide PRO templates, but every PDF endpoint takes `template`
    straight from a query string or form field, so a free user could simply
    ask for ?template=modern. The tier check has to live here, not in the UI.
    """
    if not template or template not in TEMPLATES:
        template = _user_settings(user["id"]).get("default_template", FREE_TEMPLATE)
    if template not in TEMPLATES:
        template = FREE_TEMPLATE
    limits = db.get_tier_limits(user.get("tier", "free"))
    if not limits.get("all_templates", False):
        template = FREE_TEMPLATE
    return template


def _check_tier_feature(user, feature_key):
    """Check if user's tier has a boolean feature enabled."""
    limits = db.get_tier_limits(user.get("tier", "free"))
    return bool(limits.get(feature_key, False))


def _check_tier_limit(user, resource_type):
    """Check if user has reached their tier limit for a resource type.
    Returns (allowed: bool, current_count: int, max_count: int)."""
    limits = db.get_tier_limits(user.get("tier", "free"))
    usage = db.get_user_resource_counts(user["id"])
    key_map = {
        "documents": "max_documents",
        "clients": "max_clients",
        "products": "max_products",
    }
    max_key = key_map.get(resource_type, "max_documents")
    current = usage.get(resource_type, 0)
    maximum = limits.get(max_key, 50)
    return current < maximum, current, maximum


def _get_logo_path(user_id):
    logo_dir = os.path.join(os.path.dirname(BASE_DIR), "data", "logos")
    filename = db.get_user_setting(user_id, "logo_filename")
    if filename:
        # Prevent path traversal — only allow simple filenames
        if "/" in filename or "\\" in filename or ".." in filename:
            return None
        path = os.path.join(logo_dir, filename)
        if os.path.exists(path):
            return path
    return None


def _ensure_offline_user():
    """Create or retrieve the single local user for offline mode."""
    user = db.get_user_by_username("local")
    if not user:
        db.create_user("local", "offline", display_name="Lietotājs", is_admin=True)
        user = db.get_user_by_username("local")
        if user:
            # Set tier to admin so all features are unlocked
            conn = db.get_connection()
            conn.execute("UPDATE users SET tier = 'admin' WHERE id = ?", (user["id"],))
            conn.commit()
            conn.close()
            user = db.get_user(user["id"])
    return user


class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path

        if (path.startswith("/static") or path in ("/login", "/register", "/pricing", "/contacts", "/terms", "/license",
                                                    "/robots.txt", "/sitemap.xml")
                or path == "/everypay/callback" or path == "/api/registry/search"):
            return await call_next(request)

        # Offline mode: require valid license before anything else
        if OFFLINE_MODE and not _is_offline_licensed() and path != "/license":
            return RedirectResponse("/license", status_code=303)

        # Offline mode: auto-login as local user, skip auth entirely
        if OFFLINE_MODE:
            user = _ensure_offline_user()
            if user:
                request.state.user = user
                # Skip setup redirect — let user access everything
                setup_exempt = {"/settings", "/setup", "/logout", "/set-password"}
                if path not in setup_exempt and not path.startswith("/static") and not path.startswith("/api/") and not path.startswith("/settings/logo"):
                    if not db.get_user_setting(user["id"], "company_name"):
                        return RedirectResponse("/setup", status_code=303)
                return await call_next(request)

        user = _get_current_user(request)

        if not user:
            # "/" serves a public landing page for guests
            if path == "/":
                request.state.user = None
                return await call_next(request)
            if path.startswith("/api/"):
                return JSONResponse({"error": "Unauthorized"}, status_code=401)
            return RedirectResponse("/login", status_code=303)

        if user["must_change_password"] and path != "/set-password" and path != "/logout":
            return RedirectResponse("/set-password", status_code=303)

        # Check if user needs to complete initial business setup
        setup_exempt = {"/settings", "/setup", "/logout", "/set-password"}
        if path not in setup_exempt and not path.startswith("/static") and not path.startswith("/api/") and not path.startswith("/settings/logo"):
            if not db.get_user_setting(user["id"], "company_name"):
                return RedirectResponse("/setup", status_code=303)

        request.state.user = user
        return await call_next(request)


app.add_middleware(AuthMiddleware)


def _calc_next_run(current_date, frequency):
    """Calculate the next run date based on frequency."""
    if isinstance(current_date, str):
        current_date = datetime.date.fromisoformat(current_date)
    if frequency == "monthly":
        month = current_date.month + 1
        year = current_date.year
        if month > 12:
            month = 1
            year += 1
        day = min(current_date.day, [31,29 if year%4==0 and (year%100!=0 or year%400==0) else 28,31,30,31,30,31,31,30,31,30,31][month-1])
        return datetime.date(year, month, day)
    elif frequency == "bimonthly":
        month = current_date.month + 2
        year = current_date.year
        while month > 12:
            month -= 12
            year += 1
        day = min(current_date.day, [31,29 if year%4==0 and (year%100!=0 or year%400==0) else 28,31,30,31,30,31,31,30,31,30,31][month-1])
        return datetime.date(year, month, day)
    elif frequency == "quarterly":
        month = current_date.month + 3
        year = current_date.year
        while month > 12:
            month -= 12
            year += 1
        day = min(current_date.day, [31,29 if year%4==0 and (year%100!=0 or year%400==0) else 28,31,30,31,30,31,31,30,31,30,31][month-1])
        return datetime.date(year, month, day)
    elif frequency == "halfyearly":
        month = current_date.month + 6
        year = current_date.year
        while month > 12:
            month -= 12
            year += 1
        day = min(current_date.day, [31,29 if year%4==0 and (year%100!=0 or year%400==0) else 28,31,30,31,30,31,31,30,31,30,31][month-1])
        return datetime.date(year, month, day)
    elif frequency == "yearly":
        year = current_date.year + 1
        day = min(current_date.day, [31,29 if year%4==0 and (year%100!=0 or year%400==0) else 28,31,30,31,30,31,31,30,31,30,31][current_date.month-1])
        return datetime.date(year, current_date.month, day)
    return current_date


def _fill_email_template(template: str, variables: dict) -> str:
    """Substitute {var} placeholders in an email template (subject or body)."""
    out = template or ""
    for key, value in variables.items():
        out = out.replace("{" + key + "}", str(value or ""))
    return out


def _build_email_defaults(doc, settings):
    """Return (subject, body) defaults for a document's outbound email.
    Uses the user's saved email_template from settings if set, otherwise a
    built-in fallback. Same logic is used by the doc view, dashboard, list,
    and the send route — so the user's saved template is the source of truth.
    """
    company_name = settings.get("company_name", "") or ""
    if doc.get("doc_type") == "buy":
        doc_type_name = settings.get("buy_doc_name", "Rēķins")
    elif doc.get("doc_type") == "offer":
        doc_type_name = settings.get("offer_doc_name", "Piedāvājums")
    else:
        doc_type_name = settings.get("sell_doc_name", "Rēķins")
    raw_date = doc.get("doc_date", "") or ""
    if raw_date and "-" in raw_date:
        dp = raw_date.split("-")
        display_date = f"{dp[2]}.{dp[1]}.{dp[0]}"
    else:
        display_date = raw_date

    subject = f"{company_name} - {doc_type_name} {doc['doc_number']}"
    tpl = settings.get("email_template", "") or ""
    if tpl:
        body = (tpl.replace("{doc_type}", doc_type_name)
                   .replace("{doc_number}", doc["doc_number"])
                   .replace("{date}", display_date)
                   .replace("{company}", company_name))
    else:
        body = (f"Labdien!\n\nPielikumā nosūtām dokumentu: {doc_type_name} "
                f"Nr. {doc['doc_number']}\nDatums: {display_date}\n\n"
                f"Ar cieņu,\n{company_name}\n")
    return subject, body


async def _process_recurring_invoices():
    """Background task that checks and processes due recurring invoices."""
    while True:
        try:
            today = datetime.date.today().isoformat()
            due = db.get_due_recurring_invoices(today)

            for rec in due:
                try:
                    items = json.loads(rec["items_json"])
                    if not items:
                        continue

                    # Check monthly document limit before creating
                    rec_user = db.get_user(rec["user_id"])
                    if rec_user:
                        allowed, _, _ = _check_tier_limit(rec_user, "documents")
                        if not allowed:
                            continue

                    doc_id, doc_number = db.create_document(
                        rec["user_id"], rec["doc_type"], rec["client_id"],
                        today, items, rec["vat_rate"], rec["notes"]
                    )

                    # Log document_created event (recurring source)
                    try:
                        db.log_event(rec["user_id"], "document_created",
                                     document_id=doc_id, client_id=rec["client_id"],
                                     meta={"doc_number": doc_number, "doc_type": rec["doc_type"],
                                           "source": "recurring", "recurring_id": rec["id"]})
                    except Exception:
                        logger.exception("Failed to log document_created event (recurring)")

                    template = rec.get("template", "minimal")
                    filepath = generate_invoice_pdf(doc_id, template=template)

                    # Send email if enabled
                    if rec["send_email"] and (SMTP_PASS or BREVO_API_KEY):
                        client = db.get_client(rec["client_id"])
                        if client and client.get("email"):
                            settings = db.get_all_user_settings(rec["user_id"])
                            company_name = settings.get("company_name", "")
                            doc_type_name = settings.get("sell_doc_name", "Rēķins") if rec["doc_type"] == "sell" else settings.get("buy_doc_name", "Rēķins")
                            rec_user = db.get_user(rec["user_id"])
                            user_email = rec_user.get("email", "") if rec_user else ""
                            tp = today.split("-")
                            display_date = f"{tp[2]}.{tp[1]}.{tp[0]}" if len(tp) == 3 else today

                            email_vars = {
                                "doc_type": doc_type_name,
                                "doc_number": doc_number,
                                "date": display_date,
                                "company": company_name,
                                "client_name": client.get("name", ""),
                            }
                            subject_tpl = rec.get("email_subject") or f"{doc_type_name} Nr. {{doc_number}} — {{company}}"
                            body_tpl = rec.get("email_body") or settings.get("email_template", "") or \
                                f"Labdien!\n\nPielikumā nosūtām dokumentu: {{doc_type}} Nr. {{doc_number}}\nDatums: {{date}}\n\nAr cieņu,\n{{company}}\n"

                            subject = _fill_email_template(subject_tpl, email_vars)
                            body = _fill_email_template(body_tpl, email_vars)

                            _send_email(
                                to_email=client["email"],
                                subject=subject,
                                body=body,
                                reply_to=user_email,
                                attachment_path=filepath,
                                sender_name=company_name,
                            )
                            db.log_email_sent(rec["user_id"], doc_id, client["email"], source="recurring")
                            try:
                                db.log_event(rec["user_id"], "document_sent",
                                             document_id=doc_id, client_id=rec["client_id"],
                                             meta={"recipient": client["email"], "send_type": "recurring",
                                                   "doc_number": doc_number, "recurring_id": rec["id"]})
                            except Exception:
                                logger.exception("Failed to log document_sent event (recurring)")

                    # Calculate next run
                    next_run = _calc_next_run(rec["next_run"], rec["frequency"])
                    db.update_recurring_next_run(rec["id"], next_run.isoformat())
                    logger.info(f"Recurring #{rec['id']}: created doc #{doc_id}, next run {next_run}")

                except Exception as e:
                    logger.error(f"Error processing recurring invoice #{rec['id']}: {e}")

        except Exception as e:
            logger.error(f"Error in recurring invoice loop: {e}")

        # Purge documents deleted more than 7 days ago
        try:
            db.purge_old_deleted_documents()
        except Exception as e:
            print(f"[TRASH] Purge error: {e}")

        await asyncio.sleep(3600)  # Check every hour


# --- Subscription renewals ---

MAX_RENEWAL_ATTEMPTS = 3  # After this many failures, downgrade to free


def _charge_subscription_renewal(user: dict) -> bool:
    """Attempt to renew one user's subscription. Returns True on success."""
    user_id = user["id"]
    tier = user.get("tier", "")
    cycle = user.get("billing_cycle", "monthly")
    token = user.get("everypay_token", "")
    email = user.get("email", "")

    amount = everypay.get_plan_price(tier, cycle)
    if not amount or not token:
        logger.warning("renewal: skip user %s — missing amount(%s) or token", user_id, amount)
        return False

    order_ref = f"VRR-{user_id}-{datetime.datetime.utcnow().strftime('%y%m%d%H%M%S')}"[:30]

    try:
        result = everypay.charge_mit(
            amount=amount,
            order_reference=order_ref,
            token=token,
            email=email,
        )
    except Exception:
        logger.exception("renewal: EveryPay charge failed for user %s", user_id)
        db.record_renewal_failure(user_id)
        return False

    payment_state = result.get("payment_state", "")
    payment_ref = result.get("payment_reference", "")
    logger.info("renewal: user=%s tier=%s cycle=%s amount=%.2f state=%s ref=%s",
                user_id, tier, cycle, amount, payment_state, payment_ref)

    if payment_state != "settled":
        attempts = db.record_renewal_failure(user_id)
        logger.warning("renewal: user %s charge not settled (state=%s, attempt %s/%s)",
                       user_id, payment_state, attempts, MAX_RENEWAL_ATTEMPTS)
        if attempts >= MAX_RENEWAL_ATTEMPTS:
            logger.warning("renewal: user %s exceeded max attempts, downgrading to free", user_id)
            try:
                db.cancel_user_subscription(user_id)
                _notify_renewal_final_failure(user)
            except Exception:
                logger.exception("renewal: downgrade failed for user %s", user_id)
        else:
            _notify_renewal_retry(user, attempts)
        return False

    # Settled — extend subscription, generate invoice + email
    days = 365 if cycle == "yearly" else 30
    new_end = db.extend_subscription(user_id, days, payment_ref)
    logger.info("renewal: user %s extended until %s", user_id, new_end)

    try:
        _generate_subscription_invoice(user, tier, cycle, amount, order_ref, payment_ref)
    except Exception:
        logger.exception("renewal: invoice/email generation failed for user %s", user_id)
        # Charge succeeded so don't roll back the renewal — admin can resend manually
    return True


def _notify_renewal_retry(user, attempt):
    """Email the user that we couldn't charge their card; we'll retry tomorrow."""
    email = user.get("email", "")
    if not email:
        return
    try:
        _send_email(
            to_email=email,
            subject="V-Rēķini — neizdevās atjaunot abonementu",
            body=(
                f"Labdien!\n\n"
                f"Neizdevās noīrēt maksājumu par V-Rēķini abonementa atjaunošanu (mēģinājums {attempt}/{MAX_RENEWAL_ATTEMPTS}).\n"
                f"Mēģināsim vēlreiz nākamajā dienā. Lūdzu, pārbaudiet maksājumu kartes datus sadaļā Mans konts.\n\n"
                f"Ja problēma nepāries, jūsu konts pēc {MAX_RENEWAL_ATTEMPTS} neveiksmīgiem mēģinājumiem tiks atjaunots uz bezmaksas plānu.\n\n"
                f"Ar cieņu,\nV-Rēķini\n"
            ),
            sender_name="V-Rēķini",
        )
    except Exception:
        logger.exception("renewal: retry notification email failed for user %s", user.get("id"))


def _notify_renewal_final_failure(user):
    """Email the user that we've given up and downgraded them to free."""
    email = user.get("email", "")
    if not email:
        return
    try:
        _send_email(
            to_email=email,
            subject="V-Rēķini — abonements atjaunots uz bezmaksas plānu",
            body=(
                f"Labdien!\n\n"
                f"Pēc vairākiem neveiksmīgiem maksājuma mēģinājumiem jūsu V-Rēķini abonements ir atjaunots uz bezmaksas plānu.\n"
                f"Jūsu dati ir saglabāti — varat jebkurā brīdī izvēlēties jaunu plānu sadaļā Mans abonements.\n\n"
                f"Ar cieņu,\nV-Rēķini\n"
            ),
            sender_name="V-Rēķini",
        )
    except Exception:
        logger.exception("renewal: final-failure notification email failed for user %s", user.get("id"))


async def _process_subscription_renewals():
    """Daily loop that charges saved card tokens for due subscriptions."""
    # Wait a minute after startup before first run
    await asyncio.sleep(60)
    while True:
        try:
            due = db.get_users_due_for_renewal()
            if due:
                logger.info("renewal: %d user(s) due for renewal", len(due))
            for user in due:
                try:
                    _charge_subscription_renewal(user)
                except Exception:
                    logger.exception("renewal: unexpected error for user %s", user.get("id"))
        except Exception:
            logger.exception("renewal: loop iteration failed")

        # Daily cadence
        await asyncio.sleep(24 * 3600)


@app.on_event("startup")
async def startup():
    db_path = db.get_db_path()
    db_exists = os.path.exists(db_path)
    db_size = os.path.getsize(db_path) if db_exists else 0
    print(f"[STARTUP] Database path: {db_path}")
    print(f"[STARTUP] Database exists: {db_exists}, size: {db_size} bytes")

    db.init_db()

    user_ct = db.user_count()
    print(f"[STARTUP] Users in database: {user_ct}")

    temp_pw = db.ensure_default_admin()
    if temp_pw:
        print(f"\n{'='*60}")
        print(f"  WARNING: Fresh database detected — no existing users found!")
        print(f"  Izveidots noklusējuma administrators:")
        print(f"  Lietotājvārds: admin")
        print(f"  Parole: {temp_pw}")
        print(f"  (Parole jāmaina pie pirmās pieslēgšanās)")
        print(f"{'='*60}\n")

    # Initialize registry DB table (data imported separately via cron)
    registry.init_registry_db()
    reg_count = registry.get_record_count()
    logger.info(f"[STARTUP] Business registry: {reg_count} records")
    logger.info(f"[STARTUP] EveryPay URL: {everypay.API_URL}")
    logger.info(f"[STARTUP] EveryPay account: {everypay.ACCOUNT_NAME or '(unset)'}")
    logger.info(f"[STARTUP] EveryPay configured: {everypay.is_configured()}")

    # Start recurring invoice background task
    asyncio.create_task(_process_recurring_invoices())
    # Start subscription renewal background task
    asyncio.create_task(_process_subscription_renewals())


# =============================================================================
# Offline license activation
# =============================================================================

@app.get("/license", response_class=HTMLResponse)
async def license_page(request: Request, error: str = ""):
    if not OFFLINE_MODE or _is_offline_licensed():
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse("license.html", {"request": request, "error": error})


@app.post("/license")
async def license_activate(request: Request, license_key: str = Form(...)):
    if not OFFLINE_MODE:
        return RedirectResponse("/", status_code=303)
    if _activate_license(license_key):
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse("license.html", {
        "request": request,
        "error": "Nederīga licences atslēga. Lūdzu pārbaudiet un mēģiniet vēlreiz.",
    })


# =============================================================================
# Auth routes
# =============================================================================

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, error: str = "", message: str = ""):
    user = _get_current_user(request)
    if user and not user["must_change_password"]:
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse("login.html", {
        "request": request,
        "error": error,
        "message": message,
        "page": "login",
    })


@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    # Rate limit: 10 login attempts per minute per IP
    ip = _get_client_ip(request)
    if not _rate_limiter.is_allowed(f"login:{ip}", 10, 60):
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Pārāk daudz mēģinājumu. Lūdzu, uzgaidiet.",
            "username": username,
            "page": "login",
        })

    user = db.authenticate_user(username, password)
    if not user:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Nepareizs e-pasts vai parole.",
            "username": username,
            "page": "login",
        })

    if user["must_change_password"]:
        response = RedirectResponse("/set-password", status_code=303)
    else:
        response = RedirectResponse("/", status_code=303)

    return _set_session_cookie(response, user["id"])


@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request, error: str = "", plan: str = "", cycle: str = "monthly"):
    user = _get_current_user(request)
    if user:
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse("register.html", {
        "request": request,
        "error": error,
        "page": "register",
        "plan": plan,
        "cycle": cycle,
        "entity_type": "business",
    })


@app.post("/register")
async def register(request: Request,
                   email: str = Form(...),
                   display_name: str = Form(""),
                   first_name: str = Form(""),
                   last_name: str = Form(""),
                   phone: str = Form(""),
                   password: str = Form(...),
                   confirm_password: str = Form(...),
                   plan: str = Form(""),
                   cycle: str = Form("monthly"),
                   entity_type: str = Form("business"),
                   reg_number: str = Form(""),
                   vat_number: str = Form(""),
                   legal_address: str = Form(""),
                   is_vat_payer: str = Form("0")):
    # Rate limit: 5 registrations per hour per IP
    ip = _get_client_ip(request)
    if not _rate_limiter.is_allowed(f"register:{ip}", 5, 3600):
        return templates.TemplateResponse("register.html", {
            "request": request, "page": "register", "plan": plan, "cycle": cycle,
            "entity_type": entity_type, "email": email,
            "error": "Pārāk daudz reģistrāciju. Lūdzu, mēģiniet vēlāk.",
        })

    # For individuals, build display_name from first + last name
    if entity_type == "individual":
        display_name = f"{first_name.strip()} {last_name.strip()}".strip()

    error_ctx = {
        "request": request, "page": "register",
        "email": email, "display_name": display_name, "phone": phone,
        "plan": plan, "cycle": cycle, "entity_type": entity_type,
        "first_name": first_name, "last_name": last_name,
        "reg_number": reg_number, "vat_number": vat_number,
        "legal_address": legal_address, "is_vat_payer": is_vat_payer,
    }

    if not display_name:
        err = "Lūdzu norādiet vārdu un uzvārdu." if entity_type == "individual" else "Lūdzu norādiet uzņēmuma nosaukumu."
        return templates.TemplateResponse("register.html", {**error_ctx, "error": err})

    if password != confirm_password:
        return templates.TemplateResponse("register.html", {**error_ctx, "error": "Paroles nesakrīt."})

    if len(password) < 6:
        return templates.TemplateResponse("register.html", {**error_ctx, "error": "Parolei jābūt vismaz 6 simbolus garai."})

    if db.get_user_by_email(email):
        return templates.TemplateResponse("register.html", {**error_ctx, "error": "E-pasts jau reģistrēts."})

    # Auto-generate username from email
    username = email.split("@")[0].lower().replace(" ", "_")
    base_username = username
    counter = 1
    while db.get_user_by_username(username):
        username = f"{base_username}{counter}"
        counter += 1

    user_id = db.create_user(
        username=username,
        password=password,
        display_name=display_name,
        email=email,
        phone=phone,
        tier="business" if OFFLINE_MODE else "free",
    )

    # Save registration data as settings (but NOT company_name — that's set
    # during onboarding to gate the setup-complete check in middleware)
    is_business = entity_type == "business"
    db.save_all_user_settings(user_id, {
        "invoice_number_type": "type1",
        "invoice_number_separator": "-",
        "invoice_number_digits": "3",
        "default_vat_rate": "21" if (is_business and is_vat_payer == "1") else "0",
        "is_vat_payer": is_vat_payer if is_business else "0",
        "stock_enabled": "0",
        "status_tracking": "1",
        "buy_doc_prefix": "",
        "sell_doc_prefix": "",
        "buy_doc_name": "Rēķins",
        "sell_doc_name": "Rēķins",
        "default_template": "minimal",
        "entity_type": entity_type,
        "_reg_company_name": display_name,
        "reg_number": reg_number if is_business else "",
        "vat_number": vat_number if (is_business and is_vat_payer == "1") else "",
        "legal_address": legal_address if is_business else "",
    })

    # Store pending plan selection (if any) — checkout happens after onboarding
    if plan in ("starter", "business"):
        db.save_all_user_settings(user_id, {
            "_pending_plan": plan,
            "_pending_plan_cycle": cycle,
        })

    # Always redirect to dashboard → onboarding gate will send to /setup
    response = RedirectResponse("/", status_code=303)
    return _set_session_cookie(response, user_id)


@app.get("/set-password", response_class=HTMLResponse)
async def set_password_page(request: Request, error: str = ""):
    user = _get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse("set_password.html", {
        "request": request,
        "display_name": user["display_name"] or user["username"],
        "error": error,
    })


@app.post("/set-password")
async def set_password(request: Request,
                       new_password: str = Form(...),
                       confirm_password: str = Form(...)):
    user = _get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    if new_password != confirm_password:
        return templates.TemplateResponse("set_password.html", {
            "request": request,
            "display_name": user["display_name"] or user["username"],
            "error": "Paroles nesakrīt.",
        })

    if len(new_password) < 6:
        return templates.TemplateResponse("set_password.html", {
            "request": request,
            "display_name": user["display_name"] or user["username"],
            "error": "Parolei jābūt vismaz 6 simbolus garai.",
        })

    db.update_user_password(user["id"], new_password)
    response = RedirectResponse("/", status_code=303)
    return _set_session_cookie(response, user["id"])


@app.get("/logout")
async def logout(request: Request):
    response = RedirectResponse("/login", status_code=303)
    response.delete_cookie(SESSION_COOKIE)
    return response


# =============================================================================
# Initial Setup (simplified onboarding)
# =============================================================================

@app.get("/setup", response_class=HTMLResponse)
async def setup_page(request: Request):
    user = request.state.user
    # If already set up, go to dashboard
    if db.get_user_setting(user["id"], "company_name"):
        return RedirectResponse("/", status_code=303)
    settings = db.get_all_user_settings(user["id"])
    # Use _reg_company_name from registration as the company name for pre-fill
    if not settings.get("company_name") and settings.get("_reg_company_name"):
        settings["company_name"] = settings["_reg_company_name"]
    has_logo = bool(_get_logo_path(user["id"]))
    return templates.TemplateResponse("setup.html", {
        "request": request,
        "display_name": user.get("display_name") or "",
        "settings": settings,
        "has_logo": has_logo,
        "offline_mode": OFFLINE_MODE,
    })


@app.post("/setup")
async def save_setup(
    request: Request,
    entity_type: str = Form("business"),
    company_name: str = Form(""),
    reg_number: str = Form(""),
    vat_number: str = Form(""),
    legal_address: str = Form(""),
    bank_name: str = Form(""),
    bank_account: str = Form(""),
    is_vat_payer: str = Form("0"),
    default_vat_rate: str = Form("21"),
    invoice_number_type: str = Form("type1"),
    invoice_number_separator: str = Form("-"),
    invoice_number_digits: str = Form("3"),
    electronic_doc: str = Form("0"),
    payment_due_days: str = Form(""),
):
    user = request.state.user
    settings_dict = {
        "entity_type": entity_type,
        "company_name": company_name,
        "reg_number": reg_number,
        "vat_number": vat_number,
        "legal_address": legal_address,
        "bank_name": bank_name,
        "bank_account": bank_account,
        "is_vat_payer": is_vat_payer,
        "default_vat_rate": default_vat_rate if is_vat_payer == "1" else "0",
        "invoice_number_type": invoice_number_type,
        "invoice_number_separator": invoice_number_separator,
        "invoice_number_digits": invoice_number_digits,
        "electronic_doc": electronic_doc,
        "payment_due_days": payment_due_days,
    }
    db.save_all_user_settings(user["id"], settings_dict)

    # Offline mode: already on business tier, go straight to dashboard
    if OFFLINE_MODE:
        return RedirectResponse("/", status_code=303)

    # After onboarding, check if user has a pending plan from registration
    all_settings = db.get_all_user_settings(user["id"])
    pending_plan = all_settings.get("_pending_plan", "")
    pending_cycle = all_settings.get("_pending_plan_cycle", "monthly")

    if pending_plan in ("starter", "business"):
        # Clear the pending plan and redirect to checkout
        db.save_all_user_settings(user["id"], {
            "_pending_plan": "",
            "_pending_plan_cycle": "",
        })
        return RedirectResponse(
            f"/pricing?upgrade={pending_plan}&cycle={pending_cycle}",
            status_code=303,
        )

    # No pending plan — drop them on the dashboard. Sending someone to the
    # pricing page before they have made a single invoice asks them to pay for
    # something they have not seen yet; the plans are one sidebar click away.
    return RedirectResponse("/", status_code=303)


# =============================================================================
# Account
# =============================================================================

@app.get("/account", response_class=HTMLResponse)
async def account_page(request: Request):
    ctx = _base_context(request)
    user = request.state.user
    limits = db.get_tier_limits(user.get("tier", "free"))
    usage = db.get_user_resource_counts(user["id"])
    ctx.update({
        "page": "account",
        "limits": limits,
        "usage": usage,
        "doc_count": usage["documents"],
        "has_payments": everypay.is_configured(),
        "tiers": db.TIER_LIMITS,
    })
    return templates.TemplateResponse("account.html", ctx)


@app.post("/account/profile")
async def update_profile(request: Request,
                         display_name: str = Form(...),
                         email: str = Form(""),
                         phone: str = Form("")):
    user = request.state.user
    db.update_user_profile(user["id"], display_name=display_name, email=email, phone=phone)
    return RedirectResponse("/account?saved=profile", status_code=303)


@app.post("/account/password")
async def change_password(request: Request,
                          current_password: str = Form(...),
                          new_password: str = Form(...),
                          confirm_password: str = Form(...)):
    user = request.state.user

    if not db.authenticate_user(user["username"], current_password):
        return RedirectResponse("/account?error=wrong_password", status_code=303)

    if new_password != confirm_password:
        return RedirectResponse("/account?error=mismatch", status_code=303)

    if len(new_password) < 6:
        return RedirectResponse("/account?error=too_short", status_code=303)

    db.update_user_password(user["id"], new_password)
    return RedirectResponse("/account?saved=password", status_code=303)


# =============================================================================
# Pricing (public) & Stripe billing
# =============================================================================

SITE_URL = os.getenv("VREKINI_SITE_URL", "https://v-rekini.lv")


@app.get("/robots.txt", response_class=PlainTextResponse)
async def robots_txt():
    """Allow indexing of public pages, block private/utility routes."""
    return PlainTextResponse(
        "User-agent: *\n"
        "Allow: /$\n"
        "Allow: /pricing\n"
        "Allow: /contacts\n"
        "Allow: /terms\n"
        "Disallow: /login\n"
        "Disallow: /register\n"
        "Disallow: /set-password\n"
        "Disallow: /setup\n"
        "Disallow: /account\n"
        "Disallow: /settings\n"
        "Disallow: /documents\n"
        "Disallow: /clients\n"
        "Disallow: /products\n"
        "Disallow: /recurring\n"
        "Disallow: /stock\n"
        "Disallow: /trash\n"
        "Disallow: /export\n"
        "Disallow: /email-log\n"
        "Disallow: /users\n"
        "Disallow: /billing/\n"
        "Disallow: /api/\n"
        "Disallow: /everypay/\n"
        "Disallow: /license\n"
        f"\nSitemap: {SITE_URL}/sitemap.xml\n"
    )


@app.get("/sitemap.xml")
async def sitemap_xml():
    """Sitemap listing public, indexable pages."""
    today = datetime.date.today().isoformat()
    urls = [
        ("/",         "1.0", "weekly"),
        ("/pricing",  "0.9", "weekly"),
        ("/contacts", "0.5", "monthly"),
        ("/terms",    "0.3", "yearly"),
    ]
    body = '<?xml version="1.0" encoding="UTF-8"?>\n<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    for path, prio, freq in urls:
        body += (
            "  <url>\n"
            f"    <loc>{SITE_URL}{path}</loc>\n"
            f"    <lastmod>{today}</lastmod>\n"
            f"    <changefreq>{freq}</changefreq>\n"
            f"    <priority>{prio}</priority>\n"
            "  </url>\n"
        )
    body += "</urlset>\n"
    return Response(content=body, media_type="application/xml")


@app.get("/contacts", response_class=HTMLResponse)
async def contacts_page(request: Request):
    return templates.TemplateResponse("contacts.html", {
        "request": request,
        "current_user": None,
        "page": "contacts",
    })


@app.get("/terms", response_class=HTMLResponse)
async def terms_page(request: Request):
    return templates.TemplateResponse("terms.html", {
        "request": request,
        "current_user": None,
        "page": "terms",
    })


@app.get("/pricing", response_class=HTMLResponse)
async def pricing_page(request: Request, upgrade: str = "", cycle: str = "monthly"):
    user = _get_current_user(request)
    if user:
        request.state.user = user
        ctx = _base_context(request)
        limits = db.get_tier_limits(user.get("tier", "free"))
        usage = db.get_user_resource_counts(user["id"])
        ctx.update({
            "page": "pricing",
            "limits": limits,
            "usage": usage,
            "has_payments": everypay.is_configured(),
            "upgrade": upgrade,
            "upgrade_cycle": cycle,
            "lifetime_sold": db.count_lifetime_users(),
        })
        return templates.TemplateResponse("pricing_auth.html", ctx)
    return templates.TemplateResponse("pricing.html", {
        "request": request,
        "current_user": None,
        "page": "pricing",
        "lifetime_sold": db.count_lifetime_users(),
    })


@app.post("/billing/checkout")
async def billing_checkout(request: Request, tier: str = Form(...), cycle: str = Form("monthly")):
    """Initiate an EveryPay payment and redirect to hosted payment page."""
    user = request.state.user
    if tier == "lifetime":
        cycle = "lifetime"
    if tier not in ("mini", "starter", "business", "lifetime") or cycle not in ("monthly", "yearly", "lifetime"):
        return RedirectResponse("/pricing?error=invalid", status_code=303)

    # Enforce 30-user cap for lifetime plan
    if tier == "lifetime":
        lifetime_count = db.count_lifetime_users()
        if lifetime_count >= 10:
            return RedirectResponse("/pricing?error=lifetime_sold_out", status_code=303)

    amount = everypay.get_plan_price(tier, cycle)
    if not amount or not everypay.is_configured():
        return RedirectResponse("/pricing?error=payments_not_configured", status_code=303)

    # Build order reference: VR-{plan}-{user_id}-{timestamp} (ASCII-safe, max ~20 chars for EveryPay)
    tier_code = {"mini": "M", "starter": "S", "business": "B", "lifetime": "L"}[tier]
    ts = datetime.datetime.now().strftime("%y%m%d%H%M%S")
    order_ref = f"VR{tier_code}-{user['id']}-{ts}"
    base_url = str(request.base_url).rstrip("/")
    customer_url = f"{base_url}/billing/return"
    client_ip = _get_client_ip(request)

    try:
        result = everypay.initiate_payment(
            amount=amount,
            order_reference=order_ref,
            customer_url=customer_url,
            email=user.get("email", ""),
            customer_ip=client_ip,
            request_token=True,
        )
    except Exception as e:
        logger.exception("EveryPay payment initiation failed")
        return RedirectResponse("/pricing?error=payment_error", status_code=303)

    payment_link = result.get("payment_link", "")
    payment_ref = result.get("payment_reference", "")

    if not payment_link:
        return RedirectResponse("/pricing?error=payment_error", status_code=303)

    # Store pending payment info in user settings for verification on return
    db.save_all_user_settings(user["id"], {
        "_pending_payment_ref": payment_ref,
        "_pending_order_ref": order_ref,
        "_pending_tier": tier,
        "_pending_cycle": cycle,
    })

    return RedirectResponse(payment_link, status_code=303)


def _generate_subscription_invoice(paying_user, tier, cycle, amount, order_ref, payment_reference):
    """Create a subscription invoice in the admin account and email it to the paying user.

    The invoice flows into the admin's normal document numbering and
    includes the EveryPay payment reference in the notes field.
    """
    doc_id = None
    doc_number = None
    paying_email = paying_user.get("email", "")
    try:
        # Find admin user
        all_users = db.get_all_users()
        admin = next((u for u in all_users if u.get("is_admin")), None)
        if not admin:
            logger.warning("No admin user found — skipping subscription invoice")
            return

        admin_id = admin["id"]
        today = datetime.date.today().isoformat()

        # Ensure a subscription product exists for the admin
        admin_products = db.get_all_products(admin_id)
        sub_product = next(
            (p for p in admin_products if p["name"] == "V-Rēķini abonements"), None
        )
        if not sub_product:
            sub_product_id = db.add_product(admin_id, "V-Rēķini abonements", "gab")
        else:
            sub_product_id = sub_product["id"]

        # Ensure paying user exists as a client in admin's account
        paying_settings = db.get_all_user_settings(paying_user["id"])
        client_name = (paying_settings.get("company_name", "")
                       or paying_user.get("display_name", "")
                       or paying_user["username"])
        client_reg = paying_settings.get("reg_number", "")
        client_vat = paying_settings.get("vat_number", "")

        existing_client = None
        if client_reg:
            existing_client = db.get_client_by_reg_number(admin_id, client_reg)
        if not existing_client:
            admin_clients = db.get_all_clients(admin_id)
            existing_client = next(
                (c for c in admin_clients if c["name"] == client_name), None
            )

        if existing_client:
            client_id = existing_client["id"]
        else:
            client_id = db.add_client(
                admin_id, client_name,
                reg_number=client_reg,
                vat_number=client_vat,
                vat_payer=1 if client_vat else 0,
                legal_address=paying_settings.get("legal_address", ""),
                bank_name=paying_settings.get("bank_name", ""),
                bank_account=paying_settings.get("bank_account", ""),
                email=paying_user.get("email", ""),
            )

        tier_labels = {"mini": "Mini", "starter": "Pamata", "business": "Bizness", "lifetime": "Mūža licence"}
        cycle_labels = {"monthly": "mēnesī", "yearly": "gadā", "lifetime": "vienreizējs"}
        plan_desc = f"V-Rēķini — {tier_labels.get(tier, tier)} ({cycle_labels.get(cycle, cycle)})"

        admin_settings = db.get_all_user_settings(admin_id)
        vat_rate = float(admin_settings.get("default_vat_rate", "21"))
        net_price = round(amount / (1 + vat_rate / 100), 2)

        notes = (
            f"Maksājuma ref.: {payment_reference}\n"
            f"Pasūtījuma ref.: {order_ref}"
        )

        items = [{
            "product_id": sub_product_id,
            "quantity": 1,
            "unit": "gab",
            "price_per_unit": net_price,
        }]

        doc_id, doc_number = db.create_document(
            admin_id, "sell", client_id, today, items,
            vat_rate=vat_rate, notes=notes,
        )

        try:
            db.update_document_status(admin_id, doc_id, "paid")
        except Exception:
            logger.exception("subscription_invoice: failed to mark doc %s as paid", doc_id)

        logger.info("subscription_invoice: created %s (doc_id=%s) for user %s",
                     doc_number, doc_id, paying_user.get("username", "?"))

    except Exception:
        logger.exception("subscription_invoice: failed to create invoice for user %s",
                         paying_user.get("username", "?"))
        return

    # Email step — separate try/except so a send failure doesn't hide the
    # successful invoice creation, and a send failure is visible in logs.
    if not paying_email:
        logger.warning("subscription_invoice: paying user %s has no email, skipping send",
                       paying_user.get("username", "?"))
        return

    try:
        admin_template = admin_settings.get("default_template", "minimal")
        filepath = generate_invoice_pdf(doc_id, template=admin_template)
    except Exception:
        logger.exception("subscription_invoice: PDF generation failed for doc %s", doc_id)
        return

    try:
        admin_company = admin_settings.get("company_name", "V-Rēķini")
        email_body = (
            f"Labdien!\n\n"
            f"Paldies par V-Rēķini abonementa iegādi!\n"
            f"Jūsu plāns: {plan_desc}\n"
            f"Jūsu konts ir aktīvs un gatavs lietošanai.\n\n"
            f"Pielikumā nosūtām maksājuma rēķinu Nr. {doc_number}.\n\n"
            f"Ar cieņu,\n{admin_company}\n"
        )
        _send_email(
            to_email=paying_email,
            subject=f"Rēķins Nr. {doc_number} — {admin_company}",
            body=email_body,
            attachment_path=filepath,
            sender_name=admin_company,
        )
        logger.info("subscription_invoice: emailed %s to %s", doc_number, paying_email)
    except Exception:
        logger.exception("subscription_invoice: email send failed for doc %s to %s",
                         doc_number, paying_email)


@app.get("/billing/return", response_class=HTMLResponse)
async def billing_return(request: Request, payment_reference: str = ""):
    """Handle customer return from EveryPay payment page."""
    user = request.state.user
    settings = db.get_all_user_settings(user["id"])
    pending_ref = settings.get("_pending_payment_ref", "")
    pending_tier = settings.get("_pending_tier", "")
    pending_cycle = settings.get("_pending_cycle", "")

    if not payment_reference or payment_reference != pending_ref:
        return RedirectResponse("/pricing?error=invalid_payment", status_code=303)

    try:
        status = everypay.get_payment_status(payment_reference)
    except Exception as e:
        logger.exception("Failed to check payment status")
        return RedirectResponse("/pricing?error=payment_error", status_code=303)

    payment_state = status.get("payment_state", "")

    if payment_state == "settled":
        # Extract card token for future recurring payments
        cc_details = status.get("cc_details", {})
        card_token = cc_details.get("token", "")

        # Calculate subscription end date
        if pending_cycle == "lifetime":
            end_date = "2099-12-31"
        elif pending_cycle == "yearly":
            end_date = (datetime.date.today() + datetime.timedelta(days=365)).isoformat()
        else:
            end_date = (datetime.date.today() + datetime.timedelta(days=30)).isoformat()

        # Activate subscription
        db.update_user_subscription(
            user["id"], pending_tier,
            billing_cycle=pending_cycle,
            everypay_token=card_token,
            everypay_payment_ref=payment_reference,
            subscription_status="active",
        )

        # Set subscription end date
        conn = db.get_connection()
        conn.execute("UPDATE users SET subscription_end = ? WHERE id = ?",
                     (end_date, user["id"]))
        conn.commit()
        conn.close()

        # Generate subscription invoice and email to user (idempotent — runs once)
        pending_order_ref = settings.get("_pending_order_ref", "")
        amount = everypay.get_plan_price(pending_tier, pending_cycle)
        already_invoiced = settings.get("_subscription_invoiced_ref", "") == payment_reference
        if amount and not already_invoiced:
            _generate_subscription_invoice(
                user, pending_tier, pending_cycle, amount,
                pending_order_ref, payment_reference,
            )

        # Clean up pending settings and mark invoiced
        db.save_all_user_settings(user["id"], {
            "_pending_payment_ref": "",
            "_pending_order_ref": "",
            "_pending_tier": "",
            "_pending_cycle": "",
            "_subscription_invoiced_ref": payment_reference,
        })

        logger.info("User %s upgraded to %s (%s) via EveryPay",
                     user["id"], pending_tier, pending_cycle)

        ctx = _base_context(request)
        ctx["page"] = "billing"
        return templates.TemplateResponse("billing_success.html", ctx)

    elif payment_state in ("initial", "waiting_for_3ds_response",
                           "waiting_for_sca", "sent_for_processing"):
        # Payment still processing — show waiting message
        ctx = _base_context(request)
        ctx["page"] = "billing"
        ctx["payment_pending"] = True
        ctx["payment_reference"] = payment_reference
        return templates.TemplateResponse("billing_success.html", ctx)

    else:
        # Payment failed or abandoned
        db.save_all_user_settings(user["id"], {
            "_pending_payment_ref": "",
            "_pending_order_ref": "",
            "_pending_tier": "",
            "_pending_cycle": "",
        })
        return RedirectResponse("/pricing?error=payment_failed", status_code=303)


@app.get("/everypay/callback")
async def everypay_callback(request: Request,
                            payment_reference: str = "",
                            order_reference: str = "",
                            event_name: str = ""):
    """Handle EveryPay callback notifications (server-to-server)."""
    if not payment_reference:
        raise HTTPException(status_code=400, detail="Missing payment_reference")

    logger.info("EveryPay callback: ref=%s event=%s order=%s",
                payment_reference, event_name, order_reference)

    try:
        status = everypay.get_payment_status(payment_reference)
    except Exception as e:
        logger.exception("Failed to verify payment status from callback")
        raise HTTPException(status_code=500, detail="Payment verification failed")

    payment_state = status.get("payment_state", "")

    # Look up user by pending order reference
    pending = db.find_user_by_pending_order_ref(order_reference) if order_reference else None
    if not pending:
        return JSONResponse({"status": "ok", "note": "unknown order_reference"})
    user_id = pending["user_id"]
    tier = pending["tier"]
    cycle = pending["cycle"]

    if event_name == "status_updated" and payment_state == "settled":
        cc_details = status.get("cc_details", {})
        card_token = cc_details.get("token", "")

        if cycle == "yearly":
            end_date = (datetime.date.today() + datetime.timedelta(days=365)).isoformat()
        elif cycle == "lifetime":
            end_date = "2099-12-31"
        else:
            end_date = (datetime.date.today() + datetime.timedelta(days=30)).isoformat()

        db.update_user_subscription(
            user_id, tier,
            billing_cycle=cycle,
            everypay_token=card_token,
            everypay_payment_ref=payment_reference,
            subscription_status="active",
        )

        conn = db.get_connection()
        conn.execute("UPDATE users SET subscription_end = ? WHERE id = ?",
                     (end_date, user_id))
        conn.commit()
        conn.close()

        logger.info("User %s subscription activated via callback: %s (%s)",
                     user_id, tier, cycle)

        # Generate subscription invoice + email now (idempotent on payment_reference).
        # /billing/return runs the same guard so we never double-invoice.
        paying_user = db.get_user(user_id)
        user_settings = db.get_all_user_settings(user_id)
        already_invoiced = user_settings.get("_subscription_invoiced_ref", "") == payment_reference
        amount = everypay.get_plan_price(tier, cycle)
        if paying_user and amount and not already_invoiced:
            _generate_subscription_invoice(
                paying_user, tier, cycle, amount,
                order_reference, payment_reference,
            )
            db.save_all_user_settings(user_id, {
                "_subscription_invoiced_ref": payment_reference,
            })

    elif event_name in ("abandoned", "status_updated") and payment_state == "failed":
        logger.warning("Payment failed for user %s: ref=%s", user_id, payment_reference)

    return JSONResponse({"status": "ok"})


@app.post("/billing/cancel")
async def billing_cancel(request: Request):
    """Cancel the current subscription."""
    user = request.state.user
    if user.get("tier", "free") == "free":
        return RedirectResponse("/account?error=no_subscription", status_code=303)

    # Deactivate the stored card token so no more recurring charges
    token = user.get("everypay_token", "")
    if token:
        try:
            everypay.deactivate_token(token)
        except Exception:
            logger.warning("Failed to deactivate EveryPay token for user %s", user["id"])

    db.cancel_user_subscription(user["id"])
    logger.info("User %s cancelled subscription", user["id"])
    return RedirectResponse("/account?success=subscription_cancelled", status_code=303)


# =============================================================================
# User management (admin only)
# =============================================================================

@app.get("/users", response_class=HTMLResponse)
async def users_page(request: Request, error: str = "", success: str = ""):
    ctx = _base_context(request)
    user = request.state.user
    users = db.get_all_users() if user["is_admin"] else []
    ctx.update({
        "users": users,
        "error": error,
        "success": success,
        "page": "users",
        "tiers": db.TIER_LIMITS,
    })
    return templates.TemplateResponse("users.html", ctx)


@app.post("/users/add")
async def add_user(request: Request,
                   username: str = Form(...),
                   display_name: str = Form(...),
                   password: str = Form(...),
                   email: str = Form("")):
    user = request.state.user
    if not user["is_admin"]:
        raise HTTPException(status_code=403)

    if db.get_user_by_username(username):
        return RedirectResponse("/users?error=Lietotājvārds jau aizņemts", status_code=303)

    if len(password) < 6:
        return RedirectResponse("/users?error=Parolei jābūt vismaz 6 simbolus garai", status_code=303)

    new_user_id = db.create_user(username, password, display_name, email=email,
                                 must_change_password=True)

    # Set default settings
    db.save_all_user_settings(new_user_id, {
        "invoice_number_type": "type1",
        "invoice_number_separator": "-",
        "invoice_number_digits": "3",
        "default_vat_rate": "21",
        "stock_enabled": "0",
        "status_tracking": "1",
    })

    return RedirectResponse("/users?success=Lietotājs izveidots", status_code=303)


@app.post("/users/{user_id}/delete")
async def delete_user(request: Request, user_id: int):
    user = request.state.user
    if not user["is_admin"]:
        raise HTTPException(status_code=403)
    if user_id == user["id"]:
        return RedirectResponse("/users?error=Nevar dzēst sevi", status_code=303)
    db.delete_user(user_id)
    return RedirectResponse("/users?success=Lietotājs dzēsts", status_code=303)


@app.post("/users/{user_id}/tier")
async def change_user_tier(request: Request, user_id: int, tier: str = Form(...)):
    user = request.state.user
    if not user["is_admin"]:
        raise HTTPException(status_code=403)
    if tier not in db.TIER_LIMITS:
        return RedirectResponse("/users?error=Nederīgs plāns", status_code=303)
    db.update_user_subscription(user_id, tier)
    target = db.get_user(user_id)
    name = target["username"] if target else str(user_id)
    label = db.TIER_LIMITS[tier]["label"]
    return RedirectResponse(f"/users?success=Lietotāja {name} plāns mainīts uz {label}", status_code=303)


# =============================================================================
# Dashboard
# =============================================================================

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, date_from: str = "", date_to: str = "", compare_mode: str = "auto"):
    import json as _json
    user = getattr(request.state, "user", None)
    # Guest visitor — serve public landing page (good for SEO)
    if not user:
        return templates.TemplateResponse("landing.html", {
            "request": request,
            "current_user": None,
            "page": "landing",
        })
    ctx = _base_context(request)
    uid = user["id"]
    settings = _user_settings(uid)
    stock_on = _stock_enabled(uid)

    # Default to current month
    today = datetime.date.today()
    if not date_from:
        date_from = today.replace(day=1).isoformat()
    if not date_to:
        date_to = today.isoformat()

    recent_docs = db.get_documents(uid, exclude_doc_types=["offer"])[:5]
    clients = db.get_all_clients(uid)
    stock = db.get_stock(uid) if stock_on else []
    stats = db.get_dashboard_stats(uid)
    range_stats = db.get_dashboard_stats_range(uid, date_from, date_to, compare_mode)
    recent_events = db.get_recent_events(uid, limit=5)

    ctx.update({
        "recent_docs": recent_docs,
        "clients": clients,
        "stock": stock,
        "settings": settings,
        "stats": stats,
        "range_stats": range_stats,
        "recent_events": recent_events,
        "date_from": date_from,
        "date_to": date_to,
        "range_stats_json": _json.dumps(range_stats, default=str),
        "einvoice_enabled": _check_tier_feature(user, "einvoice"),
        "recurring_enabled": _check_tier_feature(user, "recurring"),
        "email_enabled": not OFFLINE_MODE and user.get("tier", "free") != "free",
        "page": "dashboard",
    })
    return templates.TemplateResponse("dashboard.html", ctx)


@app.get("/api/dashboard-stats")
async def api_dashboard_stats(request: Request, date_from: str = "", date_to: str = "", compare_mode: str = "auto"):
    import json as _json
    user = request.state.user
    if not user:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    uid = user["id"]
    today = datetime.date.today()
    if not date_from:
        date_from = today.replace(day=1).isoformat()
    if not date_to:
        date_to = today.isoformat()
    range_stats = db.get_dashboard_stats_range(uid, date_from, date_to, compare_mode)
    return JSONResponse(range_stats)


# =============================================================================
# Settings (per-user)
# =============================================================================

@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    ctx = _base_context(request)
    user = request.state.user
    settings = _user_settings(user["id"])
    logo_path = _get_logo_path(user["id"])
    logo_v = int(os.path.getmtime(logo_path)) if logo_path else 0
    ctx.update({
        "settings": settings,
        "page": "settings",
        "has_logo": logo_path is not None,
        "logo_v": logo_v,
        "current_max_seq": db.get_user_max_seq(user["id"]),
        "templates": TEMPLATES,
    })
    return templates.TemplateResponse("settings.html", ctx)


@app.post("/settings")
async def save_settings(
    request: Request,
    entity_type: str = Form("business"),
    company_name: str = Form(""),
    reg_number: str = Form(""),
    vat_number: str = Form(""),
    legal_address: str = Form(""),
    bank_name: str = Form(""),
    bank_account: str = Form(""),
    use_prefixes: str = Form("0"),
    buy_doc_prefix: str = Form(""),
    sell_doc_prefix: str = Form(""),
    buy_doc_name: str = Form("Rēķins"),
    sell_doc_name: str = Form("Rēķins"),
    offer_doc_prefix: str = Form("P"),
    offer_doc_name: str = Form("Piedāvājums"),
    default_vat_rate: str = Form("21"),
    is_vat_payer: str = Form("0"),
    payment_due_days: str = Form(""),
    stock_enabled: str = Form("0"),
    electronic_doc: str = Form("0"),
    status_tracking: str = Form("0"),
    logo_width: str = Form("100"),
    default_template: str = Form("minimal"),
    accent_color: str = Form("#09090b"),
    invoice_number_type: str = Form("type1"),
    invoice_number_separator: str = Form(""),
    invoice_number_digits: str = Form("3"),
    invoice_number_start: str = Form("1"),
    email_template: str = Form(""),
):
    user = request.state.user
    settings_dict = {
        "entity_type": entity_type,
        "company_name": company_name,
        "reg_number": reg_number,
        "vat_number": vat_number if is_vat_payer == "1" else "",
        "legal_address": legal_address,
        "bank_name": bank_name,
        "bank_account": bank_account,
        "is_vat_payer": is_vat_payer,
        "use_prefixes": use_prefixes,
        "buy_doc_prefix": buy_doc_prefix,
        "sell_doc_prefix": sell_doc_prefix,
        "buy_doc_name": buy_doc_name,
        "sell_doc_name": sell_doc_name,
        "offer_doc_prefix": offer_doc_prefix or "P",
        "offer_doc_name": offer_doc_name or "Piedāvājums",
        "default_vat_rate": default_vat_rate if is_vat_payer == "1" else "0",
        "payment_due_days": payment_due_days,
        # Stock management is sold as a Bizness feature. Users already using it
        # keep it (turning it off under them would hide their existing data),
        # but it cannot be switched on without the tier.
        "stock_enabled": stock_enabled if (
            _check_tier_feature(user, "stock") or _stock_enabled(user["id"])
        ) else "0",
        "electronic_doc": electronic_doc,
        "status_tracking": status_tracking,
        "logo_width": logo_width,
        "default_template": default_template,
        "accent_color": accent_color if accent_color and accent_color.startswith("#") else "#09090b",
    }

    # When stock is first enabled, record the date so stock counts from 0
    if stock_enabled == "1":
        existing = db.get_user_setting(user["id"], "stock_enabled_date", "")
        if not existing:
            settings_dict["stock_enabled_date"] = datetime.date.today().isoformat()
    else:
        # If stock is disabled, clear the enabled date so re-enabling resets again
        settings_dict["stock_enabled_date"] = ""

    # Validate start number against highest existing seq for this user
    try:
        start_int = max(1, int(invoice_number_start or "1"))
    except (TypeError, ValueError):
        start_int = 1
    current_max = db.get_user_max_seq(user["id"])
    if start_int <= current_max:
        start_int = current_max + 1

    settings_dict.update({
        "invoice_number_type": invoice_number_type,
        "invoice_number_separator": invoice_number_separator,
        "invoice_number_digits": invoice_number_digits,
        "invoice_number_start": str(start_int),
        "email_template": email_template,
    })
    db.save_all_user_settings(user["id"], settings_dict)
    return RedirectResponse("/settings?saved=1", status_code=303)


@app.post("/settings/logo")
async def upload_logo(request: Request, logo: UploadFile = File(...)):
    user = request.state.user
    uid = user["id"]
    allowed = {".png", ".jpg", ".jpeg", ".gif"}
    ext = os.path.splitext(logo.filename or "")[1].lower()
    if ext not in allowed:
        return RedirectResponse("/settings?error=logo_type", status_code=303)

    logo_dir = os.path.join(os.path.dirname(BASE_DIR), "data", "logos")
    os.makedirs(logo_dir, exist_ok=True)

    # Remove old logo
    old_path = _get_logo_path(uid)
    if old_path and os.path.exists(old_path):
        os.remove(old_path)

    filename = f"{uid}_logo{ext}"
    filepath = os.path.join(logo_dir, filename)
    content = await logo.read()
    with open(filepath, "wb") as f:
        f.write(content)

    db.set_user_setting(uid, "logo_filename", filename)
    # Support redirect back to setup/onboarding
    redirect = request.query_params.get("redirect", "/settings?saved=1")
    return RedirectResponse(redirect, status_code=303)


@app.post("/settings/logo/delete")
async def delete_logo(request: Request):
    user = request.state.user
    uid = user["id"]
    path = _get_logo_path(uid)
    if path and os.path.exists(path):
        os.remove(path)
    db.set_user_setting(uid, "logo_filename", "")
    redirect = request.query_params.get("redirect", "/settings?saved=1")
    return RedirectResponse(redirect, status_code=303)


@app.get("/api/logo")
async def get_user_logo(request: Request):
    user = request.state.user
    path = _get_logo_path(user["id"])
    if path:
        return FileResponse(path, headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        })
    raise HTTPException(404)


# =============================================================================
# Products (per-user)
# =============================================================================

@app.get("/products", response_class=HTMLResponse)
async def products_page(request: Request):
    ctx = _base_context(request)
    user = request.state.user
    products = db.get_all_products(user["id"])
    usage = db.get_user_resource_counts(user["id"])
    limits = db.get_tier_limits(user.get("tier", "free"))
    ctx.update({
        "products": products,
        "units": UNITS,
        "usage": usage,
        "limits": limits,
        "page": "products",
    })
    return templates.TemplateResponse("products.html", ctx)


@app.post("/products/add")
async def add_product(request: Request, name: str = Form(...), unit: str = Form(...)):
    user = request.state.user
    allowed, current, maximum = _check_tier_limit(user, "products")
    if not allowed:
        return RedirectResponse(f"/products?error=limit", status_code=303)
    db.add_product(user["id"], name, unit)
    return RedirectResponse("/products", status_code=303)


@app.post("/products/{product_id}/edit")
async def edit_product(request: Request, product_id: int,
                       name: str = Form(...), unit: str = Form(...)):
    user = request.state.user
    db.update_product(user["id"], product_id, name, unit)
    return RedirectResponse("/products", status_code=303)


@app.post("/products/{product_id}/delete")
async def delete_product(request: Request, product_id: int):
    user = request.state.user
    db.delete_product(user["id"], product_id)
    return RedirectResponse("/products", status_code=303)


# =============================================================================
# Clients (per-user)
# =============================================================================

@app.get("/clients", response_class=HTMLResponse)
async def clients_page(request: Request):
    ctx = _base_context(request)
    user = request.state.user
    clients = db.get_all_clients(user["id"])
    usage = db.get_user_resource_counts(user["id"])
    limits = db.get_tier_limits(user.get("tier", "free"))
    ctx.update({
        "clients": clients,
        "usage": usage,
        "limits": limits,
        "page": "clients",
    })
    return templates.TemplateResponse("clients.html", ctx)


@app.post("/clients/add")
async def add_client(
    request: Request,
    name: str = Form(...),
    reg_number: str = Form(""),
    vat_number: str = Form(""),
    vat_payer: str = Form("0"),
    legal_address: str = Form(""),
    bank_name: str = Form(""),
    bank_account: str = Form(""),
    contact_person: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    client_type: str = Form("business"),
):
    user = request.state.user
    if reg_number:
        existing = db.get_client_by_reg_number(user["id"], reg_number)
        if existing:
            return RedirectResponse(f"/clients?error=duplicate&name={quote(existing['name'])}", status_code=303)
    allowed, current, maximum = _check_tier_limit(user, "clients")
    if not allowed:
        return RedirectResponse(f"/clients?error=limit", status_code=303)
    db.add_client(user["id"], name, reg_number, vat_number, int(vat_payer), legal_address,
                  bank_name, bank_account, contact_person, phone, email, client_type=client_type)
    return RedirectResponse("/clients", status_code=303)


@app.post("/clients/{client_id}/edit")
async def edit_client(
    request: Request,
    client_id: int,
    name: str = Form(...),
    reg_number: str = Form(""),
    vat_number: str = Form(""),
    vat_payer: str = Form("0"),
    legal_address: str = Form(""),
    bank_name: str = Form(""),
    bank_account: str = Form(""),
    contact_person: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    client_type: str = Form("business"),
):
    user = request.state.user
    db.update_client(user["id"], client_id, name=name, reg_number=reg_number,
                     vat_number=vat_number, vat_payer=int(vat_payer), legal_address=legal_address,
                     bank_name=bank_name, bank_account=bank_account,
                     contact_person=contact_person, phone=phone, email=email,
                     client_type=client_type)
    return RedirectResponse("/clients", status_code=303)


@app.post("/clients/{client_id}/delete")
async def delete_client(request: Request, client_id: int):
    user = request.state.user
    db.delete_client(user["id"], client_id)
    return RedirectResponse("/clients", status_code=303)


# =============================================================================
# Documents (per-user)
# =============================================================================

@app.get("/documents", response_class=HTMLResponse)
async def documents_page(request: Request, doc_type: str = "", client_id: str = "",
                         date_from: str = "", date_to: str = "", status: str = ""):
    ctx = _base_context(request)
    user = request.state.user
    docs = db.get_documents(
        user["id"],
        doc_type=doc_type or None,
        client_id=int(client_id) if client_id else None,
        date_from=date_from or None,
        date_to=date_to or None,
        status=status or None,
        exclude_doc_types=["offer"],
    )
    clients = db.get_all_clients(user["id"])
    settings = _user_settings(user["id"])
    usage = db.get_user_resource_counts(user["id"])
    limits = db.get_tier_limits(user.get("tier", "free"))
    ctx.update({
        "documents": docs,
        "clients": clients,
        "settings": settings,
        "usage": usage,
        "limits": limits,
        "einvoice_enabled": _check_tier_feature(user, "einvoice"),
        "recurring_enabled": _check_tier_feature(user, "recurring"),
        "email_enabled": not OFFLINE_MODE and user.get("tier", "free") != "free",
        "filters": {"doc_type": doc_type, "client_id": client_id,
                     "date_from": date_from, "date_to": date_to,
                     "status": status},
        "page": "documents",
    })
    return templates.TemplateResponse("documents.html", ctx)


@app.get("/offers", response_class=HTMLResponse)
async def offers_page(request: Request, client_id: str = "",
                       date_from: str = "", date_to: str = "", status: str = ""):
    """List page for Piedāvājumi — same shape as the documents list but
    scoped to doc_type='offer' only. Excluded from dashboard totals and
    from the monthly document quota."""
    ctx = _base_context(request)
    user = request.state.user
    docs = db.get_documents(
        user["id"],
        doc_type="offer",
        client_id=int(client_id) if client_id else None,
        date_from=date_from or None,
        date_to=date_to or None,
        status=status if status in OFFER_STATUSES else None,
    )
    clients = db.get_all_clients(user["id"])
    settings = _user_settings(user["id"])
    ctx.update({
        "documents": docs,
        "clients": clients,
        "settings": settings,
        "email_enabled": not OFFLINE_MODE and user.get("tier", "free") != "free",
        "filters": {"client_id": client_id, "date_from": date_from,
                    "date_to": date_to, "status": status},
        "converted_offer_ids": db.get_converted_offer_ids(user["id"]),
        "offer_statuses": OFFER_STATUSES,
        "offer_status_pill": OFFER_STATUS_PILL,
        "page": "offers",
    })
    return templates.TemplateResponse("offers.html", ctx)


@app.get("/documents/new", response_class=HTMLResponse)
async def new_document_page(request: Request, doc_type: str = "sell", from_offer: int = 0):
    ctx = _base_context(request)
    user = request.state.user
    uid = user["id"]

    # Converting an accepted offer into an invoice: load the offer and hand its
    # client, items and notes to the form as starting values. Nothing is saved
    # until the user submits, so they can still adjust quantities or prices.
    prefill_doc, prefill_items = None, []
    if from_offer:
        offer, offer_items = db.get_document(from_offer)
        if not offer or offer.get("user_id") != uid or offer.get("doc_type") != "offer":
            raise HTTPException(status_code=404, detail="Piedāvājums nav atrasts")
        prefill_doc, prefill_items = offer, offer_items
        doc_type = "sell"

    clients = _with_document_client(db.get_all_clients(uid), uid, prefill_doc)
    products = db.get_all_products(uid)
    settings = _user_settings(uid)
    stock_on = _stock_enabled(uid)
    # If stock management is off, force sell doc type
    if not stock_on and doc_type == "buy":
        return RedirectResponse("/documents/new?doc_type=sell", status_code=302)
    stock_data = db.get_stock(uid) if stock_on else []
    stock_map = {s["id"]: s["stock"] for s in stock_data}
    if doc_type == "buy":
        doc_type_name = settings.get("buy_doc_name", "Rēķins")
    elif doc_type == "offer":
        doc_type_name = settings.get("offer_doc_name", "Piedāvājums")
    else:
        doc_type_name = settings.get("sell_doc_name", "Rēķins")
    ctx.update({
        "clients": clients,
        "products": products,
        "units": UNITS,
        "settings": settings,
        "doc_type": doc_type,
        "doc_type_name": doc_type_name,
        "stock_map": stock_map,
        "templates": TEMPLATES,
        "today": datetime.date.today().isoformat(),
        "page": "new_document" if doc_type != "offer" else "new_offer",
        "doc": prefill_doc,
        "edit_items": prefill_items,
        "prefill": prefill_doc is not None,
        "from_offer": prefill_doc["id"] if prefill_doc else 0,
        "from_offer_number": prefill_doc["doc_number"] if prefill_doc else "",
    })
    return templates.TemplateResponse("document_form.html", ctx)


@app.post("/documents/create")
async def create_document(request: Request):
    user = request.state.user

    form = await request.form()
    doc_type = form.get("doc_type", "sell")

    # Tier limit applies to invoices only — offers are unlimited
    if doc_type != "offer":
        allowed, current, maximum = _check_tier_limit(user, "documents")
        if not allowed:
            return RedirectResponse(
                f"/documents?error=Ikmēneša dokumentu limits sasniegts ({current}/{maximum}). "
                f"<a href='/pricing'>Uzlabojiet plānu</a>, lai turpinātu.",
                status_code=303)

    # Block buy doc creation if stock management is off
    if doc_type == "buy" and not _stock_enabled(user["id"]):
        return RedirectResponse("/documents?error=Iegādes dokumenti nav pieejami bez noliktavas pārvaldības.", status_code=303)
    client_id = int(form.get("client_id", 0))
    doc_date = form.get("doc_date", datetime.date.today().isoformat())
    payment_due_date = form.get("payment_due_date", "")
    vat_rate = float(form.get("vat_rate", 21.0))
    reverse_charge = form.get("reverse_charge", "0") == "1"
    if reverse_charge:
        vat_rate = 0.0
    notes = form.get("notes", "")
    template = _resolve_template(user, form.get("template", ""))

    items = []
    i = 0
    while f"items[{i}][quantity]" in form:
        pid_raw = (form.get(f"items[{i}][product_id]") or "").strip()
        product_id = int(pid_raw) if pid_raw.isdigit() and pid_raw != "0" else None
        description = (form.get(f"items[{i}][product_name]") or "").strip()
        quantity = float(form[f"items[{i}][quantity]"])
        unit = form[f"items[{i}][unit]"]
        included_in_price = form.get(f"items[{i}][included_in_price]") == "1"
        price_raw = form.get(f"items[{i}][price_per_unit]") or ""
        price_per_unit = 0.0 if included_in_price or not price_raw else float(price_raw)
        # Skip blank lines (no product selected and no free-text)
        if not product_id and not description:
            i += 1
            continue
        items.append({
            "product_id": product_id,
            "description": description,
            "quantity": quantity,
            "unit": unit,
            "price_per_unit": price_per_unit,
            "included_in_price": included_in_price,
        })
        i += 1

    if not items:
        return RedirectResponse(f"/documents/new?doc_type={doc_type}&error=no_items", status_code=303)

    # Link back to the offer this invoice was built from, but only if that
    # offer really belongs to the user — the id arrives from a form field.
    from_offer_raw = (form.get("from_offer") or "").strip()
    converted_from = None
    if from_offer_raw.isdigit() and doc_type == "sell":
        src, _ = db.get_document(int(from_offer_raw))
        if src and src.get("user_id") == user["id"] and src.get("doc_type") == "offer":
            converted_from = src["id"]

    try:
        doc_id, doc_number = db.create_document(
            user["id"], doc_type, client_id, doc_date, items, vat_rate, notes,
            payment_due_date=payment_due_date, reverse_charge=reverse_charge,
            converted_from_offer_id=converted_from,
        )
    except ValueError as e:
        logger.warning("Document creation error: %s", e)
        return RedirectResponse(f"/documents/new?doc_type={doc_type}&error={quote(str(e))}", status_code=303)

    try:
        generate_invoice_pdf(doc_id, template=template)
    except Exception as e:
        logger.exception("PDF generation failed for doc %s", doc_id)

    # Invoicing an offer is the clearest possible signal the client said yes.
    if converted_from:
        try:
            db.update_document_status(user["id"], converted_from, "accepted")
        except Exception:
            logger.exception("Failed to mark offer %s accepted", converted_from)

    try:
        db.log_event(user["id"], "document_created",
                     document_id=doc_id, client_id=client_id,
                     meta={"doc_number": doc_number, "doc_type": doc_type})
    except Exception:
        logger.exception("Failed to log document_created event")

    return RedirectResponse(f"/documents/{doc_id}?created=1&template={template}", status_code=303)


@app.get("/documents/{doc_id}/edit", response_class=HTMLResponse)
async def edit_document_page(request: Request, doc_id: int):
    ctx = _base_context(request)
    user = request.state.user
    uid = user["id"]
    doc, items = db.get_document(doc_id)
    if not doc or doc.get("user_id") != uid:
        raise HTTPException(status_code=404, detail="Dokuments nav atrasts")
    clients = _with_document_client(db.get_all_clients(uid), uid, doc)
    products = db.get_all_products(uid)
    settings = _user_settings(uid)
    stock_on = _stock_enabled(uid)
    stock_data = db.get_stock(uid) if stock_on else []
    stock_map = {s["id"]: s["stock"] for s in stock_data}
    if doc["doc_type"] == "buy":
        doc_type_name = settings.get("buy_doc_name", "Rēķins")
    elif doc["doc_type"] == "offer":
        doc_type_name = settings.get("offer_doc_name", "Piedāvājums")
    else:
        doc_type_name = settings.get("sell_doc_name", "Rēķins")
    ctx.update({
        "clients": clients,
        "products": products,
        "units": UNITS,
        "settings": settings,
        "doc_type": doc["doc_type"],
        "doc_type_name": doc_type_name,
        "stock_map": stock_map,
        "templates": TEMPLATES,
        "today": datetime.date.today().isoformat(),
        "page": "offers" if doc["doc_type"] == "offer" else "documents",
        "edit_mode": True,
        "prefill": True,
        "doc": doc,
        "edit_items": items,
    })
    return templates.TemplateResponse("document_form.html", ctx)


@app.post("/documents/{doc_id}/update")
async def update_document(request: Request, doc_id: int):
    user = request.state.user
    form = await request.form()
    doc_type = form.get("doc_type", "sell")
    client_id = int(form.get("client_id", 0))
    doc_date = form.get("doc_date", datetime.date.today().isoformat())
    vat_rate = float(form.get("vat_rate", 21.0))
    reverse_charge = form.get("reverse_charge", "0") == "1"
    if reverse_charge:
        vat_rate = 0.0
    notes = form.get("notes", "")
    template = _resolve_template(user, form.get("template", ""))
    payment_due_date = form.get("payment_due_date", "")

    items = []
    i = 0
    while f"items[{i}][quantity]" in form:
        pid_raw = (form.get(f"items[{i}][product_id]") or "").strip()
        product_id = int(pid_raw) if pid_raw.isdigit() and pid_raw != "0" else None
        description = (form.get(f"items[{i}][product_name]") or "").strip()
        quantity = float(form[f"items[{i}][quantity]"])
        unit = form[f"items[{i}][unit]"]
        included_in_price = form.get(f"items[{i}][included_in_price]") == "1"
        price_raw = form.get(f"items[{i}][price_per_unit]") or ""
        price_per_unit = 0.0 if included_in_price or not price_raw else float(price_raw)
        if not product_id and not description:
            i += 1
            continue
        items.append({
            "product_id": product_id,
            "description": description,
            "quantity": quantity,
            "unit": unit,
            "price_per_unit": price_per_unit,
            "included_in_price": included_in_price,
        })
        i += 1

    if not items:
        return RedirectResponse(f"/documents/{doc_id}/edit?error=no_items", status_code=303)

    try:
        db.update_document(user["id"], doc_id, client_id, doc_date, items, vat_rate, notes,
                           payment_due_date=payment_due_date, reverse_charge=reverse_charge)
    except ValueError as e:
        logger.warning("Document update error: %s", e)
        return RedirectResponse(f"/documents/{doc_id}/edit?error={quote(str(e))}", status_code=303)

    try:
        generate_invoice_pdf(doc_id, template=template)
    except Exception as e:
        logger.exception("PDF generation failed for doc %s", doc_id)

    return RedirectResponse(f"/documents/{doc_id}?updated=1&template={template}", status_code=303)


@app.get("/documents/{doc_id}", response_class=HTMLResponse)
async def view_document(request: Request, doc_id: int, template: str = ""):
    ctx = _base_context(request)
    user = request.state.user
    doc, items = db.get_document(doc_id)
    if not doc or doc.get("user_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Dokuments nav atrasts")
    client = db.get_client(doc["client_id"])
    settings = _user_settings(user["id"])

    subtotal = sum(item["quantity"] * item["price_per_unit"] for item in items)
    vat_amount = subtotal * (doc["vat_rate"] / 100)
    total = subtotal + vat_amount

    template = _resolve_template(user, template)

    # Build default email subject + body for send modal — same logic as send
    # route, dashboard, and documents list, so the user's saved template is the
    # source of truth everywhere.
    default_subject, default_email_body = _build_email_defaults(doc, settings)

    ctx.update({
        "doc": doc,
        "items": items,
        "client": client,
        "settings": settings,
        "subtotal": subtotal,
        "vat_amount": vat_amount,
        "total": total,
        "templates": TEMPLATES,
        "selected_template": template,
        "has_logo": _get_logo_path(user["id"]) is not None,
        "default_email_subject": default_subject,
        "default_email_body": default_email_body,
        "all_templates": _check_tier_feature(user, "all_templates"),
        "einvoice_enabled": _check_tier_feature(user, "einvoice"),
        "recurring_enabled": _check_tier_feature(user, "recurring"),
        "email_enabled": not OFFLINE_MODE and user.get("tier", "free") != "free",
        "page": "offers" if doc["doc_type"] == "offer" else "documents",
        # An offer shows the invoices it produced; an invoice shows its source offer.
        "offer_invoices": (db.get_invoices_from_offer(user["id"], doc_id)
                           if doc["doc_type"] == "offer" else []),
        "offer_statuses": OFFER_STATUSES,
        "offer_status_pill": OFFER_STATUS_PILL,
        "source_offer": (db.get_document(doc["converted_from_offer_id"])[0]
                         if doc.get("converted_from_offer_id") else None),
    })
    return templates.TemplateResponse("document_view.html", ctx)


@app.get("/documents/{doc_id}/pdf")
async def download_pdf(request: Request, doc_id: int, template: str = ""):
    user = request.state.user
    doc, _ = db.get_document(doc_id)
    if not doc or doc.get("user_id") != user["id"]:
        raise HTTPException(status_code=404)
    template = _resolve_template(user, template)
    filepath = generate_invoice_pdf(doc_id, template=template)
    return FileResponse(filepath, media_type="application/pdf",
                        filename=os.path.basename(filepath))


@app.get("/api/documents/{doc_id}/email-defaults")
async def api_document_email_defaults(request: Request, doc_id: int):
    """Return the saved-template subject/body and client email for a doc.
    Used by the dashboard and documents-list send modals to pre-fill fields
    so the user sees exactly what will be sent.
    """
    user = request.state.user
    doc, _ = db.get_document(doc_id)
    if not doc or doc.get("user_id") != user["id"]:
        raise HTTPException(status_code=404)
    settings = _user_settings(user["id"])
    subject, body = _build_email_defaults(doc, settings)
    client = db.get_client(doc["client_id"]) if doc.get("client_id") else None
    return JSONResponse({
        "subject": subject,
        "body": body,
        "client_email": (client or {}).get("email", "") or "",
        "doc_number": doc.get("doc_number", ""),
    })


@app.post("/documents/{doc_id}/send")
async def send_document_email(request: Request, doc_id: int):
    """Send invoice PDF to client via email."""
    user = request.state.user
    form = await request.form()
    recipient_email = form.get("email", "").strip()
    template = _resolve_template(user, form.get("template", ""))
    return_to = form.get("return_to", "").strip()

    def _redirect_with(qs: str) -> RedirectResponse:
        """Redirect back to return_to if provided, else to the doc view."""
        if return_to:
            sep = "&" if "?" in return_to else "?"
            return RedirectResponse(f"{return_to}{sep}{qs}", status_code=303)
        return RedirectResponse(
            f"/documents/{doc_id}?{qs}&template={template}", status_code=303
        )

    doc, items = db.get_document(doc_id)
    if not doc or doc.get("user_id") != user["id"]:
        raise HTTPException(status_code=404)

    if not recipient_email:
        return _redirect_with("error=Nav norādīta e-pasta adrese")

    if not BREVO_API_KEY and not SMTP_PASS:
        return _redirect_with("error=E-pasta serviss nav konfigurēts.")

    # Check email tier limit
    limits = db.get_tier_limits(user.get("tier", "free"))
    max_emails = limits.get("max_emails_month", 0)
    if max_emails > 0:
        sent_count = db.get_emails_sent_this_month(user["id"])
        if sent_count >= max_emails:
            return _redirect_with(
                f"error=Sasniegts e-pastu limits šim mēnesim ({max_emails})."
            )
    elif not _check_tier_feature(user, "recurring"):
        if user.get("tier", "free") == "free":
            return RedirectResponse("/pricing", status_code=303)

    # Attachment: PDF by default, optionally the PEPPOL XML e-invoice.
    # XML is a paid feature and only defined for invoices, so anything that
    # doesn't qualify quietly falls back to the PDF rather than failing.
    attachment_format = (form.get("attachment_format") or "pdf").strip().lower()
    if attachment_format == "xml" and (
        doc.get("doc_type") == "offer" or not _check_tier_feature(user, "einvoice")
    ):
        attachment_format = "pdf"

    attachment_name = ""
    if attachment_format == "xml":
        try:
            # The XML lands in a temp file, so carry its real name separately
            # or the client receives something like tmpah03c_ed.xml.
            filepath, attachment_name = generate_einvoice_file(doc_id)
        except Exception as e:
            logger.exception("E-invoice XML generation failed for doc %s", doc_id)
            return _redirect_with(f"error=E-rēķina izveide neizdevās: {quote(str(e))}")
    else:
        filepath = generate_invoice_pdf(doc_id, template=template)

    # Get client and company info for email
    settings = _user_settings(user["id"])
    client = db.get_client(doc["client_id"])
    company_name = settings.get("company_name", "")
    user_email = user.get("email", "")

    # Saved-template defaults (used when the form didn't override them)
    default_subject, default_body = _build_email_defaults(doc, settings)

    custom_subject = form.get("email_subject", "").strip()
    email_subject = custom_subject if custom_subject else default_subject

    custom_body = form.get("email_body", "").strip()
    email_body = custom_body if custom_body else default_body

    # Append v-rekini.lv footer
    email_body += "\n\n---\nE-pasts sagatavots un nosūtīts no v-rekini.lv"

    try:
        _send_email(
            to_email=recipient_email,
            subject=email_subject,
            body=email_body,
            reply_to=user_email,
            attachment_path=filepath,
            attachment_name=attachment_name,
            sender_name=company_name,
        )
    except Exception as e:
        return _redirect_with(f"error=E-pasta sūtīšanas kļūda: {quote(str(e))}")

    # Log the email send
    db.log_email_sent(user["id"], doc_id, recipient_email)

    # Log activity event
    try:
        db.log_event(user["id"], "document_sent",
                     document_id=doc_id, client_id=doc.get("client_id"),
                     meta={"recipient": recipient_email, "send_type": "manual",
                           "doc_number": doc["doc_number"]})
    except Exception:
        logger.exception("Failed to log document_sent event")

    # Save the email as client's email if not already set
    if client and not client.get("email"):
        db.update_client(user["id"], client["id"], email=recipient_email)

    return _redirect_with("sent=1")


@app.post("/documents/{doc_id}/delete")
async def delete_document(request: Request, doc_id: int):
    user = request.state.user
    # Read the type before deleting so an offer sends the user back to the
    # offers list rather than dumping them in Dokumenti, where it never was.
    doc, _ = db.get_document(doc_id)
    is_offer = bool(doc) and doc.get("user_id") == user["id"] and doc.get("doc_type") == "offer"
    db.delete_document(user["id"], doc_id)
    return RedirectResponse("/offers" if is_offer else "/documents", status_code=303)


@app.get("/trash")
async def trash_page(request: Request):
    user = request.state.user
    deleted_docs = db.get_deleted_documents(user["id"])
    return templates.TemplateResponse("trash.html", {
        "request": request, "current_user": user, "page": "trash",
        "documents": deleted_docs,
        "tier": user.get("tier", "free"),
        "tier_limits": db.get_tier_limits(user.get("tier", "free")),
    })


@app.post("/trash/{doc_id}/restore")
async def restore_document(request: Request, doc_id: int):
    user = request.state.user
    db.restore_document(user["id"], doc_id)
    return RedirectResponse("/trash", status_code=303)


@app.post("/trash/{doc_id}/delete")
async def permanently_delete_document(request: Request, doc_id: int):
    user = request.state.user
    db.permanently_delete_document(user["id"], doc_id)
    return RedirectResponse("/trash", status_code=303)


@app.post("/documents/{doc_id}/status")
async def toggle_document_status(request: Request, doc_id: int):
    user = request.state.user
    doc, _ = db.get_document(doc_id)
    if not doc or doc.get("user_id") != user["id"]:
        raise HTTPException(status_code=404)
    new_status = "paid" if doc.get("status", "issued") == "issued" else "issued"
    db.update_document_status(user["id"], doc_id, new_status)
    # Return JSON for AJAX requests, fall back to redirect
    accept = request.headers.get("accept", "")
    if "application/json" in accept:
        return JSONResponse({"status": new_status})
    return RedirectResponse(f"/documents/{doc_id}", status_code=303)


@app.post("/documents/{doc_id}/offer-status")
async def set_offer_status(request: Request, doc_id: int):
    """Mark an offer as waiting / accepted / rejected."""
    user = request.state.user
    doc, _ = db.get_document(doc_id)
    if not doc or doc.get("user_id") != user["id"] or doc.get("doc_type") != "offer":
        raise HTTPException(status_code=404)
    form = await request.form()
    status = form.get("status", "")
    if status not in OFFER_STATUSES:
        raise HTTPException(status_code=400, detail="Nederīgs statuss")
    db.update_document_status(user["id"], doc_id, status)
    return RedirectResponse(f"/documents/{doc_id}", status_code=303)


@app.post("/documents/{doc_id}/toggle-stats")
async def toggle_document_stats(request: Request, doc_id: int):
    """Toggle whether a document is counted in dashboard analytics and totals."""
    user = request.state.user
    doc, _ = db.get_document(doc_id)
    if not doc or doc.get("user_id") != user["id"]:
        raise HTTPException(status_code=404)
    new_excluded = 0 if doc.get("excluded_from_stats", 0) else 1
    db.set_document_excluded(user["id"], doc_id, new_excluded)
    accept = request.headers.get("accept", "")
    if "application/json" in accept:
        return JSONResponse({"excluded_from_stats": new_excluded})
    return RedirectResponse(f"/documents/{doc_id}", status_code=303)


# =============================================================================
# Stock (per-user)
# =============================================================================

@app.get("/stock", response_class=HTMLResponse)
async def stock_page(request: Request, date_from: str = "", date_to: str = ""):
    ctx = _base_context(request)
    user = request.state.user
    stock_on = _stock_enabled(user["id"])
    stock = db.get_stock(user["id"], date_from=date_from or None,
                         date_to=date_to or None) if stock_on else []
    ctx.update({
        "stock": stock,
        "filters": {"date_from": date_from, "date_to": date_to},
        "page": "stock",
    })
    return templates.TemplateResponse("stock.html", ctx)


# =============================================================================
# Recurring Invoices
# =============================================================================

FREQUENCY_LABELS = {
    "monthly": "Katru mēnesi",
    "bimonthly": "Katrus 2 mēnešus",
    "quarterly": "Katru ceturksni",
    "halfyearly": "Katrus 6 mēnešus",
    "yearly": "Katru gadu",
}


DEFAULT_RECURRING_EMAIL_BODY = (
    "Labdien!\n\n"
    "Pielikumā nosūtām dokumentu: {doc_type} Nr. {doc_number}\n"
    "Datums: {date}\n\n"
    "Ar cieņu,\n{company}\n"
)


@app.get("/recurring", response_class=HTMLResponse)
async def recurring_page(request: Request):
    ctx = _base_context(request)
    user = request.state.user
    recurring = db.get_recurring_invoices(user["id"])
    for rec in recurring:
        try:
            items = json.loads(rec.get("items_json") or "[]")
            subtotal = sum(it.get("quantity", 0) * it.get("price_per_unit", 0) for it in items)
        except (ValueError, TypeError):
            subtotal = 0
        vat_rate = rec.get("vat_rate", 0) or 0
        rec["total"] = round(subtotal * (1 + vat_rate / 100), 2)
    ctx.update({
        "recurring": recurring,
        "frequency_labels": FREQUENCY_LABELS,
        "page": "recurring",
    })
    return templates.TemplateResponse("recurring.html", ctx)


def _render_recurring_form(request: Request, recurring=None, error: str = ""):
    """Render the create/edit recurring invoice form."""
    ctx = _base_context(request)
    user = request.state.user
    settings = _user_settings(user["id"])

    items_pre = []
    if recurring:
        try:
            raw_items = json.loads(recurring.get("items_json") or "[]")
        except (ValueError, TypeError):
            raw_items = []
        for it in raw_items:
            prod = db.get_product(it.get("product_id"))
            items_pre.append({
                "product_id": it.get("product_id"),
                "product_name": prod["name"] if prod else "",
                "unit": it.get("unit", ""),
                "quantity": it.get("quantity", 1),
                "price_per_unit": it.get("price_per_unit", 0),
            })

    default_email_body = settings.get("email_template", "") or DEFAULT_RECURRING_EMAIL_BODY
    default_email_subject = "{doc_type} Nr. {doc_number} — {company}"

    selected_client_name = ""
    selected_client_email = ""
    if recurring:
        client = db.get_client(recurring.get("client_id"))
        if client:
            selected_client_name = client.get("name", "")
            selected_client_email = client.get("email", "")

    ctx.update({
        "recurring": recurring,
        "edit_mode": recurring is not None,
        "clients": db.get_all_clients(user["id"]),
        "products": db.get_all_products(user["id"]),
        "templates": TEMPLATES,
        "frequency_labels": FREQUENCY_LABELS,
        "settings": settings,
        "units": UNITS,
        "items_pre": items_pre,
        "today": datetime.date.today().isoformat(),
        "default_email_body": default_email_body,
        "default_email_subject": default_email_subject,
        "selected_client_name": selected_client_name,
        "selected_client_email": selected_client_email,
        "error": error,
        "page": "recurring",
    })
    return templates.TemplateResponse("recurring_form.html", ctx)


@app.get("/recurring/new", response_class=HTMLResponse)
async def new_recurring_page(request: Request):
    user = request.state.user
    if not _check_tier_feature(user, "recurring"):
        return RedirectResponse("/pricing", status_code=303)
    return _render_recurring_form(request)


@app.get("/recurring/{recurring_id}/edit", response_class=HTMLResponse)
async def edit_recurring_page(request: Request, recurring_id: int):
    user = request.state.user
    if not _check_tier_feature(user, "recurring"):
        return RedirectResponse("/pricing", status_code=303)
    rec = db.get_recurring_invoice(recurring_id)
    if not rec or rec.get("user_id") != user["id"]:
        raise HTTPException(status_code=404)
    return _render_recurring_form(request, recurring=rec)


def _sync_client_email(user_id: int, client_id: int, email: str):
    """Persist the form's email back to the client record if it changed.

    Empty input does not wipe an existing email — silently ignored.
    """
    if not email:
        return
    client = db.get_client(client_id)
    if not client or client.get("user_id") != user_id:
        return
    if (client.get("email") or "").strip() != email:
        db.update_client(user_id, client_id, email=email)


def _parse_recurring_form(form):
    """Parse and validate the recurring form. Returns dict or raises ValueError."""
    try:
        client_id = int(form.get("client_id", 0))
    except (ValueError, TypeError):
        client_id = 0
    if client_id <= 0:
        raise ValueError("Lūdzu, izvēlieties klientu")

    doc_type = form.get("doc_type", "sell")
    if doc_type not in ("sell", "buy"):
        doc_type = "sell"

    try:
        vat_rate = float(form.get("vat_rate", 21.0))
    except (ValueError, TypeError):
        vat_rate = 21.0

    notes = form.get("notes", "")
    # Tier clamping happens in the callers, which have the user in scope.
    template = form.get("template", FREE_TEMPLATE)
    if template not in TEMPLATES:
        template = FREE_TEMPLATE

    frequency = form.get("frequency", "monthly")
    if frequency not in FREQUENCY_LABELS:
        frequency = "monthly"

    next_run = form.get("next_run", "").strip()
    if not next_run:
        next_run = _calc_next_run(datetime.date.today(), frequency).isoformat()
    else:
        try:
            datetime.date.fromisoformat(next_run)
        except ValueError:
            raise ValueError("Nederīgs datums pirmajai izpildei")

    send_email = form.get("send_email", "0") == "1"
    email_subject = form.get("email_subject", "").strip()
    email_body = form.get("email_body", "").strip()
    client_email = form.get("client_email", "").strip()
    if send_email and not client_email:
        raise ValueError("Lai automātiski sūtītu rēķinu, norādiet klienta e-pastu")

    items = []
    i = 0
    while f"items[{i}][product_id]" in form:
        try:
            pid = int(form[f"items[{i}][product_id]"])
            qty = float(form[f"items[{i}][quantity]"])
            price = float(form[f"items[{i}][price_per_unit]"])
        except (ValueError, TypeError):
            i += 1
            continue
        if pid > 0 and qty > 0:
            items.append({
                "product_id": pid,
                "quantity": qty,
                "unit": form[f"items[{i}][unit]"],
                "price_per_unit": price,
            })
        i += 1

    if not items:
        raise ValueError("Jāpievieno vismaz viena pozīcija")

    return {
        "doc_type": doc_type,
        "client_id": client_id,
        "vat_rate": vat_rate,
        "notes": notes,
        "template": template,
        "frequency": frequency,
        "next_run": next_run,
        "send_email": send_email,
        "email_subject": email_subject,
        "email_body": email_body,
        "client_email": client_email,
        "items": items,
    }


@app.post("/recurring/create")
async def create_recurring(request: Request):
    user = request.state.user
    if not _check_tier_feature(user, "recurring"):
        return RedirectResponse("/pricing", status_code=303)
    # Check recurring invoice limit for tier
    limits = db.get_tier_limits(user.get("tier", "free"))
    max_rec = limits.get("max_recurring", 0)
    if max_rec > 0:
        active_count = db.count_active_recurring(user["id"])
        if active_count >= max_rec:
            return RedirectResponse(f"/recurring?error=limit&max={max_rec}", status_code=303)
    form = await request.form()
    try:
        data = _parse_recurring_form(form)
    except ValueError as e:
        return _render_recurring_form(request, error=str(e))
    data["template"] = _resolve_template(user, data["template"])

    _sync_client_email(user["id"], data["client_id"], data["client_email"])

    db.create_recurring_invoice(
        user["id"], data["doc_type"], data["client_id"], data["vat_rate"],
        data["notes"], data["template"], data["frequency"], data["next_run"],
        data["send_email"], json.dumps(data["items"]),
        email_subject=data["email_subject"], email_body=data["email_body"],
    )

    return RedirectResponse("/recurring?created=1", status_code=303)


@app.post("/recurring/{recurring_id}/update")
async def update_recurring(request: Request, recurring_id: int):
    user = request.state.user
    if not _check_tier_feature(user, "recurring"):
        return RedirectResponse("/pricing", status_code=303)
    rec = db.get_recurring_invoice(recurring_id)
    if not rec or rec.get("user_id") != user["id"]:
        raise HTTPException(status_code=404)
    form = await request.form()
    try:
        data = _parse_recurring_form(form)
    except ValueError as e:
        return _render_recurring_form(request, recurring=rec, error=str(e))
    data["template"] = _resolve_template(user, data["template"])

    _sync_client_email(user["id"], data["client_id"], data["client_email"])

    db.update_recurring_invoice(
        user["id"], recurring_id, data["doc_type"], data["client_id"],
        data["vat_rate"], data["notes"], data["template"], data["frequency"],
        data["next_run"], data["send_email"], json.dumps(data["items"]),
        email_subject=data["email_subject"], email_body=data["email_body"],
    )

    return RedirectResponse("/recurring?updated=1", status_code=303)


@app.post("/recurring/from-document/{doc_id}")
async def create_recurring_from_document(request: Request, doc_id: int):
    """Create a recurring invoice schedule from an existing document."""
    user = request.state.user
    if not _check_tier_feature(user, "recurring"):
        return RedirectResponse("/pricing", status_code=303)
    # Check recurring invoice limit for tier
    limits = db.get_tier_limits(user.get("tier", "free"))
    max_rec = limits.get("max_recurring", 0)
    if max_rec > 0:
        active_count = db.count_active_recurring(user["id"])
        if active_count >= max_rec:
            return RedirectResponse(f"/recurring?error=limit&max={max_rec}", status_code=303)
    form = await request.form()
    frequency = form.get("frequency", "monthly")
    send_email = form.get("send_email", "0") == "1"
    template = _resolve_template(user, form.get("template", ""))
    next_run = form.get("next_run", "")

    doc, items = db.get_document(doc_id)
    if not doc or doc.get("user_id") != user["id"]:
        raise HTTPException(status_code=404)

    if not next_run:
        next_run = _calc_next_run(datetime.date.today(), frequency).isoformat()

    items_data = [
        {
            "product_id": item["product_id"],
            "quantity": item["quantity"],
            "unit": item["unit"],
            "price_per_unit": item["price_per_unit"],
        }
        for item in items
    ]

    db.create_recurring_invoice(
        user["id"], doc["doc_type"], doc["client_id"], doc["vat_rate"],
        doc.get("notes", ""), template, frequency, next_run, send_email,
        json.dumps(items_data)
    )

    return RedirectResponse(f"/documents/{doc_id}?scheduled=1&template={template}", status_code=303)


@app.post("/recurring/{recurring_id}/toggle")
async def toggle_recurring(request: Request, recurring_id: int):
    user = request.state.user
    if not _check_tier_feature(user, "recurring"):
        return RedirectResponse("/pricing", status_code=303)
    # Check limit when activating (not when deactivating)
    rec = db.get_recurring_invoice(recurring_id)
    if rec and not rec.get("active"):
        limits = db.get_tier_limits(user.get("tier", "free"))
        max_rec = limits.get("max_recurring", 0)
        if max_rec > 0 and db.count_active_recurring(user["id"]) >= max_rec:
            return RedirectResponse(f"/recurring?error=limit&max={max_rec}", status_code=303)
    db.toggle_recurring_invoice(user["id"], recurring_id)
    return RedirectResponse("/recurring", status_code=303)


@app.post("/recurring/{recurring_id}/delete")
async def delete_recurring(request: Request, recurring_id: int):
    user = request.state.user
    if not _check_tier_feature(user, "recurring"):
        return RedirectResponse("/pricing", status_code=303)
    db.delete_recurring_invoice(user["id"], recurring_id)
    return RedirectResponse("/recurring", status_code=303)


# =============================================================================
# Email Log
# =============================================================================

@app.get("/email-log", response_class=HTMLResponse)
async def email_log_page(request: Request):
    ctx = _base_context(request)
    user = request.state.user
    source = request.query_params.get("source")
    if source not in ("manual", "recurring"):
        source = None
    emails = db.get_email_log(user["id"], source=source)
    ctx.update({
        "emails": emails,
        "active_source": source,
        "page": "email_log",
    })
    return templates.TemplateResponse("email_log.html", ctx)


# =============================================================================
# Export
# =============================================================================

@app.get("/export", response_class=HTMLResponse)
async def export_page(request: Request):
    ctx = _base_context(request)
    user = request.state.user
    docs = db.get_documents(user["id"], exclude_doc_types=["offer"])
    settings = _user_settings(user["id"])
    # Load saved accounting export presets
    raw_presets = settings.get("accounting_export_presets", "[]")
    try:
        custom_presets = json.loads(raw_presets) if raw_presets else []
    except (json.JSONDecodeError, TypeError):
        custom_presets = []
    ctx.update({
        "documents": docs,
        "templates": TEMPLATES,
        "selected_template": settings.get("default_template", "minimal"),
        "page": "export",
        "custom_presets": custom_presets,
    })
    return templates.TemplateResponse("export.html", ctx)


@app.post("/export/pdf")
async def export_pdf_bulk(
    request: Request,
    date_from: str = Form(""),
    date_to: str = Form(""),
    doc_type: str = Form(""),
    template: str = Form("minimal"),
):
    import zipfile
    import tempfile

    user = request.state.user
    template = _resolve_template(user, template)
    docs = db.get_documents(
        user["id"],
        doc_type=doc_type or None,
        date_from=date_from or None,
        date_to=date_to or None,
        exclude_doc_types=["offer"],
    )

    if not docs:
        return RedirectResponse("/export?error=no_docs", status_code=303)

    # Create ZIP with all PDFs
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    tmp.close()

    with zipfile.ZipFile(tmp.name, "w", zipfile.ZIP_DEFLATED) as zf:
        for doc in docs:
            filepath = generate_invoice_pdf(doc["id"], template=template)
            arcname = f"{doc['doc_number'].replace('/', '_')}.pdf"
            zf.write(filepath, arcname)

    filename = f"dokumenti_{date_from}_{date_to}.zip"
    return FileResponse(
        tmp.name,
        media_type="application/zip",
        filename=filename,
    )


# =============================================================================
# E-invoice XML Export (PEPPOL BIS Billing 3.0)
# =============================================================================

@app.get("/documents/{doc_id}/einvoice")
async def download_einvoice(request: Request, doc_id: int):
    """Download a single document as PEPPOL BIS 3.0 e-invoice XML."""
    user = request.state.user
    if not _check_tier_feature(user, "einvoice"):
        return RedirectResponse("/pricing", status_code=303)
    doc, _ = db.get_document(doc_id)
    if not doc or doc["user_id"] != user["id"]:
        raise HTTPException(status_code=404)
    filepath, filename = generate_einvoice_file(doc_id)
    return FileResponse(
        filepath,
        media_type="application/xml",
        filename=filename,
    )


@app.post("/export/einvoice")
async def export_einvoice_bulk(
    request: Request,
    date_from: str = Form(""),
    date_to: str = Form(""),
    doc_type: str = Form(""),
):
    """Export multiple documents as e-invoice XML files in a ZIP archive."""
    import zipfile
    import tempfile

    user = request.state.user
    if not _check_tier_feature(user, "einvoice"):
        return RedirectResponse("/pricing", status_code=303)
    docs = db.get_documents(
        user["id"],
        doc_type=doc_type or None,
        date_from=date_from or None,
        date_to=date_to or None,
        exclude_doc_types=["offer"],
    )

    if not docs:
        return RedirectResponse("/export?error=no_docs", status_code=303)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    tmp.close()

    with zipfile.ZipFile(tmp.name, "w", zipfile.ZIP_DEFLATED) as zf:
        for doc in docs:
            xml_string, filename = generate_einvoice_xml(doc["id"])
            zf.writestr(filename, xml_string)

    zip_filename = f"e-rekini_{date_from}_{date_to}.zip"
    return FileResponse(
        tmp.name,
        media_type="application/zip",
        filename=zip_filename,
    )


# =============================================================================
# Accounting Export (Grāmatvedības eksports)
# =============================================================================

# Built-in presets — placeholder column configs (will be refined with real specs)
ACCOUNTING_PRESETS = {
    "horizon": {
        "name": "Horizon",
        "doc_columns": [
            {"header": "Dokumenta tips", "source": "doc_type_code"},
            {"header": "Klients.Kods", "source": "client_reg_number"},
            {"header": "PVN kategorija", "source": "vat_category"},
            {"header": "Dok. datums", "source": "doc_date", "format": "dd.mm.yyyy"},
            {"header": "Valūta.Kods", "source": "constant", "value": "EUR"},
            {"header": "Numurs", "source": "doc_number"},
            {"header": "Summa bez PVN", "source": "subtotal_no_vat"},
            {"header": "PVN summa", "source": "vat_amount"},
            {"header": "Summa ar PVN", "source": "total_with_vat"},
            {"header": "Apmaksas termiņš", "source": "payment_due_date", "format": "dd.mm.yyyy"},
            {"header": "Klients", "source": "client_name"},
            {"header": "Klients.PVN", "source": "client_vat_number"},
            {"header": "Adrese", "source": "client_address"},
            {"header": "E-pasts", "source": "client_email"},
        ],
        "item_columns": [
            {"header": "Pavadzīme.Numurs", "source": "doc_number"},
            {"header": "Nosaukums", "source": "product_name"},
            {"header": "Tips", "source": "constant", "value": "3"},
            {"header": "Daudzums", "source": "quantity"},
            {"header": "Summa", "source": "line_total"},
        ],
    },
    "jumis": {
        "name": "Jumis",
        "doc_columns": [
            {"header": "Dok.tips", "source": "doc_type_code"},
            {"header": "Numurs", "source": "doc_number"},
            {"header": "Datums", "source": "doc_date", "format": "dd.mm.yyyy"},
            {"header": "Klients", "source": "client_name"},
            {"header": "Reģ.Nr.", "source": "client_reg_number"},
            {"header": "PVN Nr.", "source": "client_vat_number"},
            {"header": "Summa bez PVN", "source": "subtotal_no_vat"},
            {"header": "PVN", "source": "vat_amount"},
            {"header": "Kopā", "source": "total_with_vat"},
            {"header": "Valūta", "source": "constant", "value": "EUR"},
        ],
        "item_columns": [
            {"header": "Dok.numurs", "source": "doc_number"},
            {"header": "Prece", "source": "product_name"},
            {"header": "Mērv.", "source": "product_unit"},
            {"header": "Daudzums", "source": "quantity"},
            {"header": "Cena", "source": "price_per_unit"},
            {"header": "Summa", "source": "line_total"},
        ],
    },
    "zalktis": {
        "name": "Zalktis",
        "doc_columns": [
            {"header": "Tips", "source": "doc_type_code"},
            {"header": "Nr.", "source": "doc_number"},
            {"header": "Datums", "source": "doc_date", "format": "dd.mm.yyyy"},
            {"header": "Partneris", "source": "client_name"},
            {"header": "Reģ.nr.", "source": "client_reg_number"},
            {"header": "PVN nr.", "source": "client_vat_number"},
            {"header": "Adrese", "source": "client_address"},
            {"header": "Bez PVN", "source": "subtotal_no_vat"},
            {"header": "PVN", "source": "vat_amount"},
            {"header": "Ar PVN", "source": "total_with_vat"},
        ],
        "item_columns": [
            {"header": "Dok.nr.", "source": "doc_number"},
            {"header": "Prece/pakalpojums", "source": "product_name"},
            {"header": "Mērvienība", "source": "product_unit"},
            {"header": "Daudzums", "source": "quantity"},
            {"header": "Cena bez PVN", "source": "price_per_unit"},
            {"header": "Summa bez PVN", "source": "line_total"},
        ],
    },
}

# All available data fields for column builder
ACCOUNTING_EXPORT_FIELDS = {
    "doc_type_code": "Dokumenta tips (Pirk./Pārd.)",
    "doc_type_name": "Dokumenta tips (nosaukums)",
    "doc_number": "Dokumenta numurs",
    "doc_date": "Datums",
    "payment_due_date": "Apmaksas datums",
    "seq_num": "Secības numurs",
    "vat_rate": "PVN likme (%)",
    "vat_category": "PVN kategorija (M/X)",
    "notes": "Piezīmes",
    "status": "Statuss",
    "subtotal_no_vat": "Summa bez PVN",
    "vat_amount": "PVN summa",
    "total_with_vat": "Summa ar PVN",
    "client_name": "Klients — nosaukums",
    "client_reg_number": "Klients — reģ. nr.",
    "client_vat_number": "Klients — PVN nr.",
    "client_address": "Klients — adrese",
    "client_bank": "Klients — banka",
    "client_account": "Klients — konts",
    "client_contact": "Klients — kontaktpersona",
    "client_phone": "Klients — telefons",
    "client_email": "Klients — e-pasts",
    "company_name": "Uzņēmums — nosaukums",
    "company_reg": "Uzņēmums — reģ. nr.",
    "company_vat": "Uzņēmums — PVN nr.",
    "company_address": "Uzņēmums — adrese",
    "company_bank": "Uzņēmums — banka",
    "company_account": "Uzņēmums — konts",
    "product_name": "Prece — nosaukums",
    "product_unit": "Prece — mērvienība",
    "quantity": "Daudzums",
    "price_per_unit": "Cena par vienību",
    "line_total": "Rindas summa",
    "constant": "Fiksēta vērtība",
}

# Fields only available in line-item sheet
_ITEM_ONLY_FIELDS = {"product_name", "product_unit", "quantity", "price_per_unit", "line_total"}


def _resolve_field_value(source, doc, item, settings, fmt=None, constant_val=""):
    """Resolve a column field value from document/item/settings data."""
    val = ""
    if source == "constant":
        val = constant_val
    elif source == "doc_type_code":
        val = "Pirk." if doc.get("doc_type") == "buy" else "Pārd."
    elif source == "doc_type_name":
        val = "Pārdošana" if doc.get("doc_type") == "sell" else "Iegāde"
    elif source == "doc_number":
        val = doc.get("doc_number", "")
    elif source == "doc_date":
        raw = doc.get("doc_date", "")
        if fmt == "dd.mm.yyyy" and raw:
            parts = raw.split("-")
            if len(parts) == 3:
                val = f"{parts[2]}.{parts[1]}.{parts[0]}"
            else:
                val = raw
        else:
            val = raw
    elif source == "payment_due_date":
        raw = doc.get("payment_due_date", "")
        if fmt == "dd.mm.yyyy" and raw:
            parts = raw.split("-")
            if len(parts) == 3:
                val = f"{parts[2]}.{parts[1]}.{parts[0]}"
            else:
                val = raw
        else:
            val = raw
    elif source == "seq_num":
        val = doc.get("seq_num", "")
    elif source == "vat_rate":
        val = doc.get("vat_rate", 21.0)
    elif source == "vat_category":
        vp = doc.get("client_vat_payer", 0)
        val = "M" if vp and str(vp) != "0" else "X"
    elif source == "notes":
        val = doc.get("notes", "") or ""
    elif source == "status":
        val = doc.get("status", "") or ""
    elif source in ("subtotal_no_vat", "vat_amount", "total_with_vat"):
        # Calculated from items
        items = doc.get("items", [])
        subtotal = sum(i["quantity"] * i["price_per_unit"] for i in items)
        vat_rate = doc.get("vat_rate", 21.0) or 21.0
        vat_amount = subtotal * (vat_rate / 100)
        if source == "subtotal_no_vat":
            val = round(subtotal, 2)
        elif source == "vat_amount":
            val = round(vat_amount, 2)
        else:
            val = round(subtotal + vat_amount, 2)
    elif source.startswith("client_"):
        val = doc.get(source, "") or ""
    elif source.startswith("company_"):
        key_map = {
            "company_name": "company_name",
            "company_reg": "reg_number",
            "company_vat": "vat_number",
            "company_address": "legal_address",
            "company_bank": "bank_name",
            "company_account": "bank_account",
        }
        val = settings.get(key_map.get(source, ""), "") or ""
    elif source == "product_name":
        val = (item or {}).get("product_name", "")
    elif source == "product_unit":
        val = (item or {}).get("product_unit", "") or (item or {}).get("unit", "")
    elif source == "quantity":
        val = (item or {}).get("quantity", "")
    elif source == "price_per_unit":
        val = (item or {}).get("price_per_unit", "")
    elif source == "line_total":
        val = (item or {}).get("total", "")
    return val


@app.post("/export/accounting")
async def export_accounting(request: Request):
    """Generate accounting Excel export with 2 sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    import tempfile

    user = request.state.user
    if not _check_tier_feature(user, "accounting_export"):
        return RedirectResponse("/pricing", status_code=303)
    form = await request.form()
    date_from = form.get("acc_date_from", "")
    date_to = form.get("acc_date_to", "")
    doc_type = form.get("acc_doc_type", "")
    preset_key = form.get("acc_preset", "")

    settings = _user_settings(user["id"])

    # Load preset config
    if preset_key in ACCOUNTING_PRESETS:
        preset = ACCOUNTING_PRESETS[preset_key]
    elif preset_key.startswith("custom_"):
        # Load from user's saved presets
        raw = settings.get("accounting_export_presets", "[]")
        try:
            custom_presets = json.loads(raw) if raw else []
        except (json.JSONDecodeError, TypeError):
            custom_presets = []
        idx = int(preset_key.replace("custom_", ""))
        if 0 <= idx < len(custom_presets):
            preset = custom_presets[idx]
        else:
            return RedirectResponse("/export?error=no_preset", status_code=303)
    else:
        return RedirectResponse("/export?error=no_preset", status_code=303)

    docs = db.get_documents_for_export(
        user["id"],
        doc_type=doc_type or None,
        date_from=date_from or None,
        date_to=date_to or None,
    )

    if not docs:
        return RedirectResponse("/export?error=no_docs", status_code=303)

    doc_columns = preset.get("doc_columns", [])
    item_columns = preset.get("item_columns", [])

    wb = Workbook()

    # --- Sheet 1: Dokumenti (one row per document) ---
    ws_docs = wb.active
    ws_docs.title = "Dokumenti"

    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Headers
    for col_idx, col_def in enumerate(doc_columns, 1):
        cell = ws_docs.cell(row=1, column=col_idx, value=col_def.get("header", ""))
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    for row_idx, doc in enumerate(docs, 2):
        for col_idx, col_def in enumerate(doc_columns, 1):
            val = _resolve_field_value(
                col_def.get("source", ""),
                doc, None, settings,
                fmt=col_def.get("format"),
                constant_val=col_def.get("value", ""),
            )
            cell = ws_docs.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-width columns
    for col_idx, col_def in enumerate(doc_columns, 1):
        max_len = len(col_def.get("header", ""))
        for row in ws_docs.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws_docs.column_dimensions[ws_docs.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)

    # --- Sheet 2: Pozīcijas (one row per line item) ---
    if item_columns:
        ws_items = wb.create_sheet("Pozīcijas")

        for col_idx, col_def in enumerate(item_columns, 1):
            cell = ws_items.cell(row=1, column=col_idx, value=col_def.get("header", ""))
            cell.font = header_font
            cell.alignment = header_alignment

        row_idx = 2
        for doc in docs:
            for item in doc.get("items", []):
                for col_idx, col_def in enumerate(item_columns, 1):
                    val = _resolve_field_value(
                        col_def.get("source", ""),
                        doc, item, settings,
                        fmt=col_def.get("format"),
                        constant_val=col_def.get("value", ""),
                    )
                    cell = ws_items.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                row_idx += 1

        for col_idx, col_def in enumerate(item_columns, 1):
            max_len = len(col_def.get("header", ""))
            for row in ws_items.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws_items.column_dimensions[ws_items.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)

    # Save and return
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)

    preset_name = preset.get("name", "eksports").lower().replace(" ", "_")
    filename = f"gramatvediba_{preset_name}_{date_from}_{date_to}.xlsx"
    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@app.post("/api/accounting-presets/save")
async def save_accounting_presets(request: Request):
    """Save custom accounting export presets."""
    user = request.state.user
    data = await request.json()
    presets = data.get("presets", [])
    db.set_user_setting(user["id"], "accounting_export_presets", json.dumps(presets))
    return JSONResponse({"ok": True})


@app.get("/api/accounting-presets")
async def get_accounting_presets(request: Request):
    """Get built-in and custom presets."""
    user = request.state.user
    settings = _user_settings(user["id"])
    raw = settings.get("accounting_export_presets", "[]")
    try:
        custom = json.loads(raw) if raw else []
    except (json.JSONDecodeError, TypeError):
        custom = []
    return JSONResponse({
        "builtin": {k: {"name": v["name"], "doc_columns": v["doc_columns"], "item_columns": v["item_columns"]} for k, v in ACCOUNTING_PRESETS.items()},
        "custom": custom,
        "fields": ACCOUNTING_EXPORT_FIELDS,
        "item_only_fields": list(_ITEM_ONLY_FIELDS),
    })


# =============================================================================
# API endpoints for AJAX
# =============================================================================

@app.post("/api/products/add")
async def api_add_product(request: Request):
    user = request.state.user
    allowed, current, maximum = _check_tier_limit(user, "products")
    if not allowed:
        return JSONResponse({"error": f"Produktu limits sasniegts ({current}/{maximum})"}, status_code=403)
    data = await request.json()
    product_id = db.add_product(user["id"], data["name"], data["unit"])
    product = db.get_product(product_id)
    return JSONResponse(product)


@app.get("/api/documents")
async def api_documents(request: Request, doc_type: str = "", client_id: str = "",
                        date_from: str = "", date_to: str = "", status: str = ""):
    user = request.state.user
    docs = db.get_documents(
        user["id"],
        doc_type=doc_type or None,
        client_id=int(client_id) if client_id else None,
        date_from=date_from or None,
        date_to=date_to or None,
        status=status or None,
        exclude_doc_types=["offer"],
    )
    settings = _user_settings(user["id"])
    status_tracking = settings.get("status_tracking", "0") == "1"
    rows = []
    for d in docs:
        rows.append({
            "id": d["id"],
            "doc_number": d["doc_number"],
            "doc_type": d["doc_type"],
            "client_name": d["client_name"],
            "client_id": d.get("client_id", 0),
            "doc_date": d["doc_date"],
            "total_with_vat": round(d.get("total_with_vat") or 0, 2),
            "status": d.get("status", "issued") if status_tracking else None,
            "excluded_from_stats": d.get("excluded_from_stats", 0),
        })
    return JSONResponse(rows)


@app.post("/api/clients/add")
async def api_add_client(request: Request):
    user = request.state.user
    data = await request.json()
    one_time = int(data.get("one_time", 0))
    reg_number = data.get("reg_number", "")
    # One-time clients are attached to a single document and never appear in
    # the clients list, so they don't trigger duplicate checks or count
    # against the tier quota.
    if not one_time:
        if reg_number:
            existing = db.get_client_by_reg_number(user["id"], reg_number)
            if existing:
                return JSONResponse({"error": f"Klients ar reģ. nr. {reg_number} jau eksistē ({existing['name']})", "duplicate": True, "client": existing}, status_code=409)
        allowed, current, maximum = _check_tier_limit(user, "clients")
        if not allowed:
            return JSONResponse({"error": f"Klientu limits sasniegts ({current}/{maximum})"}, status_code=403)
    client_id = db.add_client(
        user["id"],
        name=data["name"],
        reg_number=data.get("reg_number", ""),
        vat_number=data.get("vat_number", ""),
        vat_payer=int(data.get("vat_payer", 0)),
        legal_address=data.get("legal_address", ""),
        bank_name=data.get("bank_name", ""),
        bank_account=data.get("bank_account", ""),
        client_type=data.get("client_type", "business"),
        one_time=one_time,
    )
    client = db.get_client(client_id)
    return JSONResponse(client)


@app.get("/api/stock/{product_id}")
async def api_product_stock(request: Request, product_id: int):
    user = request.state.user
    stock = db.get_product_stock(user["id"], product_id)
    return JSONResponse({"stock": stock})


@app.get("/api/invoice-preview")
async def api_invoice_preview(request: Request,
                              number_type: str = "type1",
                              separator: str = "-",
                              digits: str = "3"):
    """Preview what invoice numbers will look like."""
    today = datetime.date.today()
    year_short = today.year % 100
    min_digits = int(digits) if digits.isdigit() else 3

    if number_type == "type1":
        example = f"{year_short}{separator}{'1'.zfill(min_digits)}"
        example2 = f"{year_short}{separator}{'42'.zfill(min_digits)}"
    elif number_type == "type2":
        day = today.day
        month = today.month
        example = f"01/{day:02d}-{month:02d}"
        example2 = f"05/{day:02d}-{month:02d}"
    else:
        example = "1".zfill(min_digits)
        example2 = "42".zfill(min_digits)

    return JSONResponse({"example1": example, "example2": example2})


# ---------- Business Registry Search ----------

@app.get("/api/registry/search")
async def api_registry_search(request: Request, q: str = ""):
    """Search the Latvian business registry by name or registration number.
    Public endpoint — registry data is already publicly available."""
    # Rate limit: 30 requests per minute per IP
    ip = _get_client_ip(request)
    if not _rate_limiter.is_allowed(f"registry:{ip}", 30, 60):
        return JSONResponse({"error": "Rate limited"}, status_code=429)
    results = registry.search(q, limit=15)
    return JSONResponse(results)


@app.get("/api/registry/status")
async def api_registry_status(request: Request):
    """Check if the business registry database is loaded."""
    user = request.state.user
    if not user or not user.get("is_admin"):
        return JSONResponse({"error": "Unauthorized"}, status_code=403)
    count = registry.get_record_count()
    return JSONResponse({"count": count, "loaded": count > 0})


# ---------- VIES VAT Number Validation ----------

@app.get("/api/vat/validate")
async def api_vat_validate(request: Request, vat_number: str = ""):
    """Validate an EU VAT number via the VIES SOAP service."""
    user = request.state.user
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    vat_number = vat_number.strip().replace(" ", "").replace("-", "")
    if len(vat_number) < 4:
        return JSONResponse({"valid": False, "error": "PVN numurs par īsu"})

    # Extract country code and number
    country_code = vat_number[:2].upper()
    number = vat_number[2:]

    # If no country prefix, assume Latvia
    if country_code.isdigit():
        country_code = "LV"
        number = vat_number

    try:
        import httpx
        soap_body = f"""<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
                  xmlns:urn="urn:ec.europa.eu:taxud:vies:services:checkVat:types">
    <soapenv:Body>
        <urn:checkVat>
            <urn:countryCode>{country_code}</urn:countryCode>
            <urn:vatNumber>{number}</urn:vatNumber>
        </urn:checkVat>
    </soapenv:Body>
</soapenv:Envelope>"""

        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.post(
                "https://ec.europa.eu/taxation_customs/vies/services/checkVatService",
                content=soap_body,
                headers={"Content-Type": "text/xml; charset=utf-8"},
            )

        body = resp.text
        valid = "<valid>true</valid>" in body.lower() or "<ns2:valid>true</ns2:valid>" in body.lower()

        # Extract name and address from response
        name = ""
        address = ""
        for tag in ["name", "ns2:name"]:
            start = body.find(f"<{tag}>")
            end = body.find(f"</{tag}>")
            if start != -1 and end != -1:
                name = body[start + len(tag) + 2:end].strip()
                break
        for tag in ["address", "ns2:address"]:
            start = body.find(f"<{tag}>")
            end = body.find(f"</{tag}>")
            if start != -1 and end != -1:
                address = body[start + len(tag) + 2:end].strip()
                break

        return JSONResponse({
            "valid": valid,
            "country_code": country_code,
            "vat_number": number,
            "name": name,
            "address": address,
        })

    except Exception as e:
        logger.error("VIES validation error: %s", e)
        return JSONResponse({"valid": False, "error": "VIES serviss nav pieejams"})
